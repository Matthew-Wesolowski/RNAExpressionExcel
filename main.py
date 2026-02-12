# ver DEV 02/12/2026

import pandas as pd
from dataclasses import dataclass
import string
import numpy as np
import scipy.stats
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
import pickle
import re
from alive_progress import alive_bar
import multiprocessing as mtp
from multiprocessing import Queue
from multiprocessing import Pool
from multiprocessing import Process, Manager
import time
import traceback
import tkinter as tk
from tkinter import filedialog
from tkinter import *
from tkinter.ttk import *
from bs4 import BeautifulSoup
import requests
import urllib.parse
import warnings
import time
import io
from io import StringIO
from PIL import ImageGrab, ImageFont, Image, ImageDraw
import mygene
from copy import deepcopy
import openpyxl
import umap
import math
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
import seaborn as sns
#from sksurv.nonparametric import kaplan_meier_estimator
from lifelines import KaplanMeierFitter
from lifelines.statistics import logrank_test
from lifelines.datasets import load_waltons
import os
import GEOparse
from goatools.anno.idtogos_reader import IdToGosReader
from goatools.go_enrichment import GOEnrichmentStudy
from goatools.obo_parser import GODag
import randomcolor

import anndata as ad

import pdb

# use a dictionary and have a text file with defined values
# TODO: This code is almost a thousand lines long, try to split it up into multiple files... <-- this first I guess

# TODO: heatmap of fischer transform for ulcerated vs non ulcerated
# TODO: More functionality! Class to interpret TCGA barcode, include sequencing and methylated regions, etc.

# TODO: train a classifier to segregate the N/A ulceration data
# TODO: Rework saving to account for duplicates!
# TODO: Include full barcode in each expression container (analayte)

# TODO: Come up with some sort of way to organize your code & save data w/ versioning dummy

# TODO: Come up with a system to make user input for expression keys easier, I don't want to have to
#       look up the keys in the original data every time I reprocess

# TODO: Isolate contexts in which mir-196b is highly expressed and has poor outcomes (ulceration, PFI, DSS)

'''
    *.*.*:
    first number contains major changes to the saving and data processing that cannot be migrated without reprocessing the original data
    second number means changes to the objects being saved and restored and is incompatable with previous binaries, meaning the data has to be migrated in order to be used with the newest version
    third number means any changes to data processing, user interfacing, and other changes that do not affect compatability with save binaries. Data can be moved seamlessly between these versions without migration
'''
TCGA_PATIENT_ID_START = 0
TCGA_PATIENT_ID_END = 12

MIR_PREDICTION_ALGORITHM = 'MIR_PREDICTION_ALGORITHM'

def FixPatientKeys():
    global TCGA_Patients
    
    with alive_bar(len(TCGA_Patients)) as bar:
        for pt in TCGA_Patients:
            for key,exp in pt.ExpressionData.items():
                if len(exp.ExpressionValues) == 1:
                    temp = list(exp.ExpressionValues.keys())[0]
                    exp.ExpressionValues['normalized_count'] = list(exp.ExpressionValues.values())[0]

                    del exp.ExpressionValues[temp]
            bar()

class TCGA_ExpressionData:
    def __init__(self):
        self.gene_ID = ''
        #self.TCGA_Barcode: string = ''
        self.ExpressionValues = {}

    def CheckExpressionType(self, exp):
        if exp in self.ExpressionValues.keys():
            return True
        return False
'''
how do we classify data for each patient?
   * overview
       ** mortality
       ** dx's
       ** melanoma type
       ** melanoma site(s)
    * Clinical follow ups
'''

class ClinicalData:
    def __init__(self, alivestatus_a, dx_a, pfi_a, survival_a):
        self.alivestatus = alivestatus_a
        self.dx = dx_a # how to normalize so its consistent accross datasets
        self.dx_site = dx_a
        self.pfi = pfi_a
        self.survival = survival_a

        #self.follow_ups <-- deal with this later


def _HandleMergeExpressionConflicts(gene : str, classexpr : str, expr : str) -> None:
    print(f"A merge conflict was detected for gene: {gene}, for expression values: {classexpr}, {expr} (calling class expr is first)")

class TCGA_Patient:
    def __init__(self, ID_a):
        self.ID = ID_a
        self.ExpressionData = {}
        self.ClinicalFeatures = {}

        self.EpigeneticData = {}
    
    def AppendClinicalFeature(self, feature, data):
        if feature not in self.ClinicalFeatures.keys():
            self.ClinicalFeatures[feature] = data
    
    # stores the expression of a certain gene for this patient
    def StoreExpression(self, df, gene_ID):
        if (gene_ID == 'miRNA_ID'):
            # this is kinda cheaty but should work
            return
        
        #get all columns with name, and then search for our gene of interest, if its not there, error and move on
        ''' Its a bit ugly, but it works, the  boolean array selects the first row so that we can have
            the type of expression data for the dictionary'''
        if gene_ID in self.ExpressionData.keys():
            # already have record of this gene, complain to the user
            print(f'Gene: {gene} for patient ID: {self.ID} is already present with expression data, how would you like to handle the conflict?')
            return # return for now since this doesn't seem to be a problem

        genedata = df.filter(regex=self.ID)
        if (genedata.empty):
            return

        exprtypeidx = df.index.get_loc(1) # expression type MUST to have loc of 0
        exparr = np.array(([False] * (len(df))))
        exparr[exprtypeidx] = True
                
        genedata = genedata.loc[(df.iloc[:,0] == gene_ID) | exparr]
        if genedata.isnull().values.any() or len(genedata.index) < 2:
            return
            ''' TODO: temporary,ideally you would want
                to still store the other values
                THIS IS TEMPORARY PLS CHANGE EVENTUALLY'''
        
        data = TCGA_ExpressionData()
        data.gene_ID = gene_ID

        # get row with data
        dataloc = df[df['Hybridization'] == gene_ID].index[0]
        
        for val in range(len(genedata.columns)):
            temp = genedata.iloc[:,val]
            data.ExpressionValues[temp.loc[1]] = temp.loc[dataloc]
            
        self.ExpressionData[gene_ID] = data
    
    def GetExpressionValue(self, gene_ID, expression):
        if self.CheckGeneDict(gene_ID, expression):
            return self.ExpressionData[gene_ID].ExpressionValues[expression]
        return None

    def CheckGeneDict(self, gene_ID, expressiontype):
        if gene_ID in self.ExpressionData.keys():
            if self.ExpressionData[gene_ID].CheckExpressionType(expressiontype):
                return True
        return False

    def GetClinicalFeature(self, feature):
        if feature in self.ClinicalFeatures.keys():
            return self.ClinicalFeatures[feature]

        return None
    
    def RemoveClinicalFeature(self, feature):
        if feature in self.ClinicalFeatures.keys():
            del self.ClinicalFeatures[feature]
    
    #combines data in dictionary
    def MergePatientExpressionData(self, pt):
        # union operator has first argument as priority for duplicates, which is what we want because technically, the values should be the same...
        merged = self.ExpressionData | pt.ExpressionData

        # include only duplicate values in sets
        for k in set(self.ExpressionData.keys()) & set(pt.ExpressionData.keys()):
            # if there are duplicates, check for conflicts
            if self.ExpressionData[k] != pt.ExpressionData[k]:
                _HandleMergeExpressionConflicts(k, self.ExpressionData[k], pt.ExpressionData[k])
                print("Abandoning merge!")

        self.ExpressionData = merged

    def MergePatientFeatureData(self, pt):
        # union operator has first argument as priority for duplicates, which is what we want because technically, the values should be the same...
        merged = self.ClinicalFeatures | pt.ClinicalFeatures

        # include only duplicate values in sets
        for k in set(self.ClinicalFeatures.keys()) & set(pt.ClinicalFeatures.keys()):
            # if there are duplicates, check for conflicts
            if self.ClinicalFeatures[k] != pt.ClinicalFeatures[k]:
                _HandleMergeExpressionConflicts(k, self.ClinicalFeatures[k], pt.ClinicalFeatures[k])
                print("Abandoning merge!")

        self.ClinicalFeatures = merged

    def MergePatientData(self, pt):
        self.MergePatientExpressionData(pt)
        self.MergePatientFeatureData(pt)

TCGA_Patients = []
GeneList = [] # list of all the stored genes
ProcessedDataBuffer = []
GeneDict = {}

# ughhh, barcode tells us if the passed value is formatted as an analayte barcode, or a patient ID
def CheckPatientDuplicates(TCGA_Barcode, ptlist = None, barcode=True):
    global TCGA_Patients

    if ptlist is None:
        ptlist = TCGA_Patients
    
    # checks for patient duplicates in array
    if barcode:
        for value in ptlist:
            if value.ID == GetPatientID(TCGA_Barcode):
                return True
    else:
        for value in ptlist:
            if value.ID == TCGA_Barcode:
                return True
    return False

# create TCGA patients off of a dataset
# TODO: other functions to add or remove patients
def CreateTCGAPatients(df, ptlist = None):
    # for now just use second row from top for names
    #namelist = df.loc[[0]].squeeze()
    #namelist = namelist.iloc[3:]
    namelist = df.columns[2:]
    
    for value in namelist:
        if (not CheckPatientDuplicates(value, ptlist=ptlist)):
            if ptlist is None:
                TCGA_Patients.append( TCGA_Patient( GetPatientID(value) ) )
            else:
                ptlist.append( TCGA_Patient( GetPatientID(value) ) )
            
def GetPatient(ID, ptlist=None):
    global TCGA_Patients

    if ptlist is None:
        ptlist = TCGA_Patients
    
    for value in TCGA_Patients:
        if (value.ID == ID):
            return value
    return None # TODO: this will cause an error eventually, keep sharp eyes

def GetPatientID(TCGA_Barcode):
    return TCGA_Barcode[TCGA_PATIENT_ID_START:TCGA_PATIENT_ID_END]    

def StoreGeneForAllPatients(df, gene_ID):
    for value in TCGA_Patients:
        value.StoreExpression(df, gene_ID)
    
    GeneList.append(gene_ID)

def StorePatientDatas(pts, ptlist=None):
    global TCGA_Patients

    if ptlist is None:
        ptlist = TCGA_Patients
    
    # check if there are duplicates
    for pt in pts:
        value = GetPatient(pt.ID,ptlist)
        if value is not None:
            GetPatient(pt.ID).MergePatientData(pt)
        else:
            ptlist.append(pt)
            
# I don't like this method but atp I don't have a choice :/
def __CreateGeneListFromPatients():
    global GeneLis
    GeneList.clear()

    with alive_bar(len(TCGA_Patients)) as bar:
        for pt in TCGA_Patients:
            for val in pt.ExpressionData.keys():
                if val not in GeneList:
                    GeneList.append(val)
            bar()

#TODO: tk window to show current gene being stored
def SubStoreEverything(df, counter, lock, ptlist):
    templist = []
    CreateTCGAPatients(df, templist)
    for value in df.iloc[:,0]:
        #print(value)
        for pt in templist:
            pt.StoreExpression(df, value)

        lock.acquire()
        counter.value += 1
        lock.release()

    lock.acquire()
    StorePatientDatas(templist, ptlist)
    lock.release()

def UI_Updater(barcounters, progress_counters, lock, win, totalitems):
    # update progress counter
    lock.acquire()
    totalcounter = 0
    for idx,bc in enumerate(barcounters):
        #print(bc.value)
        progress_counters[idx].set(bc.value)
        totalcounter += bc.value

    lock.release()
    if totalcounter >= totalitems:
        win.quit()
        win.destroy()
        return

    win.update_idletasks()
    win.after(100, UI_Updater, barcounters, progress_counters, lock, win, totalitems)

# TODO: user exit behavior

def StoreEverything(numprocs = 4):
    if numprocs > os.cpu_count():
        print(f"Input {numprocs} core count when cpu count is {os.cpucount()}!")
        return
    elif numprocs > os.cpu_count() / 2:
        print(f"Core count {numprocs} is greater than half of cpu count!")

    df = ImportTCGAData()
    
    #df = ImportTCGAData().iloc[:50] # temporary for fast processing for dev
    
    print(f"Number of Processes: {numprocs}")
    
    tmprow = df.iloc[0]
    
    df = df.iloc[1:]
    
    values = df.iloc[1:,0].tolist()

    print("Loaded file")
    
    # split everything up into numproc dfs
    
    factor = len(df.index) // numprocs

    subdfs = []
    for i in range(0,numprocs-1):
        subdfs.append(df.iloc[i*factor:(i+1)*factor].copy())
        subdfs[i].loc[1] = tmprow
        print(subdfs[i])
        
    subdfs.append(df.iloc[factor*(numprocs-1):].copy())
    subdfs[-1].loc[1] = tmprow
    print(subdfs[-1])
    
    print("Divided data")
    
    with Manager() as manager:
        barcounters = [mtp.Value('i', 0) for idx in range(numprocs)]
        
        ptlist = manager.list()
        lock = mtp.Lock()


        procs = [Process(target=SubStoreEverything, args=(s,barcounters[idx],lock,ptlist,)) for idx,s in enumerate(subdfs)]
        #procs = [Process(target=SubStoreEverything, args=(s,barcounter,lock,ptlist,)) for s in subdfs]

        print("Created Processes")
        
        # create TKwindow
        root = tk.Tk()
        root.title("Progress")

        # create progressbar for each
        progress_counters = [tk.DoubleVar() for idx in range(numprocs)]
        progress_bars = [Progressbar(root, orient='horizontal', mode='determinate', variable=progress_counters[idx], maximum=len(s.index)) for idx,s in enumerate(subdfs)]
        
        for pb in progress_bars:
            pb.pack(pady=10, padx=20)
        
        #UIProc = Process(target=UI_Updater, args=(barcounters, progress_counters, lock, root,))
        print("Created UI")

        print("Running Processes...")  
        for p in procs:
            p.start()

        print("Finished")
        #UIProc.start()

        totalitems = len(df.iloc[1:,0])

        root.after(10, UI_Updater, barcounters, progress_counters, lock, root, totalitems)

        print("Running UI...")

        root.mainloop()

        print("Finished Processing!")
        
        #with alive_bar(len(df.iloc[1:,0])) as bar:
        #localcounter = [0] * len(barcounters)
        '''totalcoutner = 0
        while(True):
            # check for change in counter
            lock.acquire()

            #for idx,bc in enumerate(barcounters):
             #   if bc.value > localcounter:
             #       while(localcounter != barcounter.value):
              #          #bar()
              #          localcounter[idx] += 1
               #         progress_counters[idx] += 1
            
            totalcounter = 0
            for idx,bc in enumerate(barcounters):
                progress_counters[idx].set(bc.value)
                totalcoutner += bc.value
            
            lock.release()

            if totalcounter == len(df.iloc[1:,0]):
                break

            root.update_idletasks()
            time.sleep(0.5)'''
        
        print("Waiting on worker procs...")
        for p in procs:
            p.join()

        print("Loading into memory...")
        StorePatientDatas(ptlist)

        print("done.")


def GetAnnData():
    # get ann data object patient list
    global TCGA_Patients
    

    


def PrintPatients():
    for value in TCGA_Patients:
        print(value.ID)

def RestorePatientDatas():
    global TCGA_Patients
    global GeneList

    
    TCGA_Patients.clear()

    root_win = Tk()
    root_win.withdraw()

    filenames = filedialog.askopenfilenames()

    tempTCGA = []
    tempGeneList = []
    tempProcessedData = []
    
    for file in filenames:
        with open(file, 'rb') as f:
            tempTCGA, tempGeneList, tempProcessedData = pickle.load(f)
            StorePatientDatas(tempTCGA)
            GeneList = GeneList + list(set(GeneList) - set(tempGeneList))

def ImportTCGAData():
    filename = filedialog.askopenfilename()
    
    df = pd.read_csv(filename, low_memory=False)
    df.columns = df.iloc[0]
    df = df[1:]
    
    return df

def IsGeneStored(gene_ID):
    #returns true if a gene is in gene list
    if gene_ID in GeneList:
        return True
    
    return False

# "legacy" code implies that this implementation should be on its way out...
# TODO: finish the new save/restore implementation
def SaveLegacy():
    global TCGA_Patients
    global GeneList
    global ProcessedDataBuffer
    
    root_win = tk.Tk() # TODO: fix so only need to create root window once
    root_win.withdraw()

    filename = filedialog.askopenfilename()
    with open(filename, 'wb') as f:
        pickle.dump((TCGA_Patients, GeneList, ProcessedDataBuffer), f)

def RestoreLegacy():
    global GeneList
    global ProcessedDataBuffer
    global TCGA_Patients

    root_win = Tk()
    root_win.withdraw()

    filename = filedialog.askopenfilename()
    
    f = open(filename, 'rb')
    tempPT, tempGL, tempPDB = pickle.load(f)
    print(tempGL)
    # todo: remove
    '''if type(tempPT[0]) is list:
        tempPT = [pt for ptlist in tempPT for pt in ptlist]
    if type(tempGL[0]) is list:
        tempGL = [pt for ptlist in tempGL for pt in ptlist]
    '''
    if len(TCGA_Patients) != 0:
        # check if we should combine the data
        StorePatientDatas(tempPT)
        GeneList = GeneList + list(set(tempGL) - set(GeneList))
    else:
        TCGA_Patients = tempPT
        GeneList = tempGL

    ProcessedDataBuffer.clear()
    ProcessedDataBuffer = tempPDB
    
    
def CombineLegacyData():
    global GeneList
    global ProcessedDataBuffer
    global TCGA_Patients

    GeneList.clear()
    ProcessedDataBuffer.clear()
    TCGA_Patients.clear()

    filenames = filedialog.askopenfilenames()
    for file in filenames:
        tempGL = None
        tempPDB = None
        tempTCGA = None

        with open(file, 'rb') as f:
            tempTCGA, tempGL, tempPDB = pickle.load(f)

        GeneList += tempGL
        ProcessedDataBuffer += tempPDB
        TCGA_Patients += tempTCGA
    
# TODO: change to work as zip file
def SavePatientData():
    root_win = tk.Tk() # TODO: fix so only need to create root window once
    root_win.withdraw()

    filename = filedialog.asksaveasfilename()
    
    with open(filename, 'wb') as f:
        pickle.dump((TCGA_Patients, GeneList), f)

def RestorePatientData():
    global TCGA_Patients
    
    GeneList.clear()
    TCGA_Patients.clear()
    ProcessedDataBuffer.clear()

    root_win = Tk()
    root_win.withdraw()

    filename = filedialog.askopenfilename()
    
    f = open(filename, 'rb')
    TCGA_Patients = pickle.load(f)

def SaveProcessedData():
    root_win = tk.Tk() # TODO: fix so only need to create root window once
    root_win.withdraw()

    filename = filedialog.asksaveasfilename()
    
    with open(filename, 'wb') as f:
        pickle.dump((GeneList, ProcessedDataBuffer), f)

def RestoreProcessedData():
    global ProcessedDataBuffer
    global GeneList
    
    GeneList.clear()
    ProcessedDataBuffer.clear()

    root_win = Tk()
    root_win.withdraw()

    filename = filedialog.askopenfilename()
    
    f = open(filename, 'rb')
    GeneList, ProcessedDataBuffer = pickle.load(f)


def RestoreGeneDict():
    global GeneDict

    root_win = Tk()
    root_win.withdraw()
    
    filename = filedialog.askopenfilename()

    f = open(filename, 'rb')
    GeneDict = pickle.load(f)

def SaveGeneDict():
    global GeneDict

    root_win = Tk()
    root_win.withdraw()

    filename = filedialog.asksaveasfilename()

    f = open(filename, 'wb')
    pickle.write(GeneDict, f)

def SaveTCGAPatients():
    global TCGA_Patients
    
    root_win = Tk()
    root_win.withdraw()
    
    filename = filedialog.asksaveasfilename()

    f = open(filename, 'wb')
    pickle.dump(TCGA_Patients, f)

def RestoreTCGAPatients():
    global TCGA_Patients
    
    root_win = Tk()
    root_win.withdraw()
    
    filename = filedialog.askopenfilename()

    f = open(filename, 'rb')
    TCGA_Patients = pickle.load(f)

class SaveDataClass:
    def __init__(self, genelist, processedbuff, patients):
        self.genelist = genelist
        self.processedbuff = processedbuff
        self.patients = patients

# for development in interpreter
def GetSaveData():
    global GeneList
    global ProcessedDataBuffer
    global TCGA_Patients

    return deepcopy( SaveDataClass(GeneList, ProcessedDataBuffer, TCGA_Patients) )

def RestoreSaveData(data):
    global GeneList
    global ProcessedDataBuffer
    global TCGA_Patients

    GeneList = data.genelist
    ProcessedDataBuffer = data.processedbuff
    TCGA_Patients = data.patients

# workspace is a compressed cluster of files that contains data on patients, processed regression data, and other things that can be retrieved using the component system class
#def OpenWorkSpace():
    # import gene list and 

# there is a lot of data to save and restore, so this component system is needed to make things easier to manage

# functor class for save and restore functions
@dataclass
class SaveRestore:
    def __init(self, savefn, restorefn):
        # save and restore should be functions
        self.savefn = savefn
        self.restorefn = restorefn

#ComponentList = {'Patient Data' : SaveRestore(), 'Gene List', 'Processed Data', 'miR Targets'}
class ComponentSystem:
    def __init__(self):
        self.Components = []

#    def Restore(self, component):
        
    
# costly, maybe include this in save file
def GetAllGenes():
    tempgenelist = []
    with alive_bar(len(TCGA_Patients)) as bar:
        for value in TCGA_Patients:
            for genekey in value.ExpressionData: # find a way to vectorize this
                if genekey not in tempgenelist:
                    tempgenelist.append(genekey)
            bar()
    return tempgenelist

def CreateGeneList():
    global GeneList
    GeneList = GetAllGenes()

def CreateGeneListFromDF(df):
    for val in df.iloc[1:0,0]:
        if val not in GeneList:
            GeneList.append(val)

def TranslateGenes(genelist):
    # converts gene names into readable gene names
    for i in range(len(genelist)):
        genelist[i] = GeneDict[genelist[i]]

class ExpressionData:
    def __init__(self, Igene, Iexpressiontype, Dgene, Dexpressiontype):
        self.IGene = Igene
        self.IGeneEtype = Iexpressiontype
        self.DGene = Dgene
        self.DGeneEtype = Dexpressiontype
        self.IValues = []
        self.DValues = []

        self.CorCoeff = 0.0
        self.pval = 0.0
        self.slope = 0.0
        self.intercept = 0.0

        # number of matches against miR prediction algorithms, etc.
        self.metadata = {}
        self.metadata[MIR_PREDICTION_ALGORITHM] = []

    def GetMetadata(self, key):
        if key in self.metadata.keys():
            return self.metadata[key]

        return None

    def ProcessData(self):
        # go through each patient, calculate the log of the independent and dependent genes and store them
        for value in TCGA_Patients:
            '''#todo: figure out a better way to do this (group genes based off of the way their expression is quantified perhaps?)
            if value.CheckGeneDict(self.IGene, self.IGeneEtype) and value.CheckGeneDict(self.DGene, self.DGeneEtype):
                Itemp = float(value.GetExpressionValue(self.IGene, self.IGeneEtype))
                Dtemp = float(value.GetExpressionValue(self.DGene, self.DGeneEtype))
                if not Itemp == 0 and not Dtemp == 0:
                    self.IValues.append(np.log2(Itemp))
                    self.DValues.append(np.log2(Dtemp))'''
            self.ProcessPatient(value)

        if (len(self.IValues) == 0 or len(self.DValues) == 0):
            return
        
        tempresult = scipy.stats.linregress(self.IValues, self.DValues)
        self.CorCoeff = tempresult.rvalue
        self.pval = tempresult.pvalue
        self.slope = tempresult.slope
        self.intercept = tempresult.intercept

    def ProcessDataForClinicalFeature(self, feature, val):
        for value in filter(lambda x: x.GetClinicalFeature(feature) == val, TCGA_Patients):
            self.ProcessPatient(value)

        if (len(self.IValues) == 0 or len(self.DValues) == 0):
            return

        tempresult = scipy.stats.linregress(self.IValues, self.DValues)
        self.CorCoeff = tempresult.rvalue
        self.pval = tempresult.pvalue
        self.slope = tempresult.slope
        self.intercept = tempresult.intercept

        self.metadata[feature] = val
        
    def ProcessPatient(self, patient):
        #todo: figure out a better way to do this (group genes based off of the way their expression is quantified perhaps?)
        if patient.CheckGeneDict(self.IGene, self.IGeneEtype) and patient.CheckGeneDict(self.DGene, self.DGeneEtype):
            Itemp = float(patient.GetExpressionValue(self.IGene, self.IGeneEtype))
            Dtemp = float(patient.GetExpressionValue(self.DGene, self.DGeneEtype))
            if not Itemp == 0 and not Dtemp == 0:
                self.IValues.append(np.log2(Itemp))
                self.DValues.append(np.log2(Dtemp))
                
    
    def PlotData(self):
        fig, ax = plt.subplots()
        
        ax.scatter(self.IValues, self.DValues)

        xlowbound = min(self.IValues)
        xhighbound = max(self.IValues)
        
        xseq = np.linspace(xlowbound, xhighbound, num=100)
        ax.plot(xseq, self.intercept + self.slope * xseq, color="k", lw=2.5)
        plt.xlabel("Log2 of expression of " + self.IGene)
        plt.ylabel("Log2 of expression of " + self.DGene)
        plt.subplots_adjust(top=0.833)

        fig.text(0.3, 0.9, "Correlation Coefficient: " + str(self.CorCoeff))
        fig.text(0.3, 0.87, "P-Value: " + str(self.pval))
        
        fig.show()
        fig.suptitle(str(self.IGene) + " vs. " + str(self.DGene))
        print('---- ' + str(self.IGene) + " vs. " + str(self.DGene) + ' ----')
        print("Correlation Coefficient: " + str(self.CorCoeff))
        print("p-value: " + str(self.pval))
        print("Regression line slope: " + str(self.slope))
        print("Regression line intercept: " + str(self.intercept))

    def PlotDataForEmbedded(self, fig, ax):
        plt.xlabel("Log2 of expression of " + self.IGene)
        plt.ylabel("Log2 of expression of " + self.DGene)
        
        ax.scatter(self.IValues, self.DValues)

        if len(self.IValues) == 0 or len(self.DValues) == 0:
            return

        try: # for some reason empty lvalue and dvalue arrays are making it to this point and then crashing the script...
            xlowbound = min(self.IValues)
            xhighbound = max(self.IValues)
        except:
            return

        fig.text(0.3, 0.9, "Correlation Coefficient: " + str(self.CorCoeff))
        fig.text(0.3, 0.87, "P-Value: " + str(self.pval))
        
        fig.suptitle(str(self.IGene) + " vs. " + str(self.DGene))
        xseq = np.linspace(xlowbound, xhighbound, num=100)
        ax.plot(xseq, self.intercept + self.slope * xseq, color="k", lw=2.5)
        plt.subplots_adjust(top=0.833)

        
    def GetSubplots(self):
        fig, ax = plt.subplots()
        
        ax.scatter(self.IValues, self.DValues)
        
        xlowbound = min(self.IValues)
        xhighbound = max(self.IValues)
        
        xseq = np.linspace(xlowbound, xhighbound, num=100)
        ax.plot(xseq, self.intercept + self.slope * xseq, color="k", lw=2.5)
        plt.xlabel("Log2 of expression of " + self.IGene)
        plt.ylabel("Log2 of expression of " + self.DGene)
        plt.subplots_adjust(top=0.833)

        return (fig,ax)

    # getters for sort functions and the like
    @staticmethod
    def GetCC(graph):
        return graph.CorCoeff

    @staticmethod
    def GetPval(graph):
        return graph.pval

    @staticmethod
    def GetSlope(graph):
        return graph.slope

    @staticmethod
    def GetIntercept(graph):
        return graph.intercept


def ProcessGenesForGeneList(Igene, Iexpressiontype, Dgenelist, Dexpressionlist):
    graphlist = []
    for idx,Dgene in enumerate(Dgenelist):
        graph = ExpressionData(Igene, Iexpressiontype, Dgene, Dexpressionlist[idx])
        graph.ProcessData()
        graphlist.append(graph)

    return graphlist

def ProcessGeneListForClinicalFeature(Igene, Iexpressiontype, Dgenelist, Dexpressionlist, feature):
    graphlist = []
    # first iterate through patient data, gather possible clinical feature values
    featurestates = []
    for pt in TCGA_Patients:
        temp = pt.GetClinicalFeature(feature)
        if temp is None:
            continue

        if type(temp) is not str:
            continue #sometimes pandas will be dumb and replace NA with NAN even when I want NA to mean 'not applicable'
        
        if temp not in featurestates:
            featurestates.append(temp)
    
    # for each possible value, compute expression data of that group
    with alive_bar(len(Dgenelist) * len(featurestates)) as bar:
        for featureval in featurestates:
            for idx,val in enumerate(Dgenelist):
                if Igene is val: continue
                tempgraph = ExpressionData(Igene, Iexpressiontype, val, Dexpressionlist[idx])
                tempgraph.ProcessDataForClinicalFeature(feature, featureval)
                graphlist.append(tempgraph)
                bar()

    return graphlist

def FastGraphGenes(Igene, Iexpressiontype, Dgene, Dexpressiontype):
    graph = ExpressionData(Igene, Iexpressiontype, Dgene, Dexpressiontype)
    graph.ProcessData()
    graph.PlotData()

# TODO: this is very much not optimal, this functionality should be housed within the ExpressionDataContainer class
def _ExpressionDataVisualizerNavigateRight(container, data, root):
    if data[0] == len(container.Data)-1:
        return
    data[0] = data[0] + 1 # index of expression data, list box data structure, figure, ax, toolbar
        
    #embed matplotlib window
    data[3].clear()
    container.Data[data[0]].PlotDataForEmbedded(data[2], data[3])
    data[2].canvas.draw()

    data[1].delete(0, END)

    data[1].insert(END, "Correlation Coefficient: " + str(container.Data[data[0]].CorCoeff))
    data[1].insert(END, "p-value: " + str(container.Data[data[0]].pval))
    data[1].insert(END, "Regression line slope: " + str(container.Data[data[0]].slope))
    data[1].insert(END, "Regression line intercept: " + str(container.Data[data[0]].intercept))
        
    for key, value in container.Data[data[0]].metadata.items():
        data[1].insert(END, key + ": " + str(value))

    data[1].insert(END, str(data[0]+1) + "/" + str(len(container.Data)))
    
    root.update()
        
def _ExpressionDataVisualizerNavigateLeft(container, data, root):
    if data[0] == 0:
        return
    data[0] = data[0] - 1 # index of expression data, list box data structure, figure, ax, toolbar
        
    #embed matplotlib window
    data[3].clear()
    container.Data[data[0]].PlotDataForEmbedded(data[2], data[3])
    data[2].canvas.draw()
    
    data[1].delete(0, END)

    data[1].insert(END, "Correlation Coefficient: " + str(container.Data[data[0]].CorCoeff))
    data[1].insert(END, "p-value: " + str(container.Data[data[0]].pval))
    data[1].insert(END, "Regression line slope: " + str(container.Data[data[0]].slope))
    data[1].insert(END, "Regression line intercept: " + str(container.Data[data[0]].intercept))
        
    for key, value in container.Data[data[0]].metadata.items():
        data[1].insert(END, key + ": " + str(value))
    
    data[1].insert(END, str(data[0]+1) + "/" + str(len(container.Data)))
    
    root.update()

def Util_WrapText(text, maxlen, font, fontsize):
    font = ImageFont.truetype(font, fontsize)
    index=0
    returnVal=''
    while True:
        temp = text.find(' ', index, len(text)-1)
        if temp == -1:
            if (index <len(text)):
                temp = len(text)-1
                # there is probably one more word
            else:
                break
        
        word = text[index:temp+1]

        tempword = returnVal + word
        tempnl = tempword.rfind('\n')+1
        if font.getlength(tempword[tempnl:]) > maxlen:
            returnVal = returnVal + '\n' + word
        else:
            returnVal = returnVal + word

        index = temp+1

    return returnVal

def _SaveWindow(window, fig, data, filepath):
    # going to have to paste graph onto new image, then write the rest of the metadata onto the library
    FONT = "arial.ttf"
    FONTSIZE = 10
    
    if filepath is None:
        return

    #draw call for tostring_rgb()
    fig.canvas.draw()
    
    buf = fig.canvas.tostring_argb()
    ncols, nrows = fig.canvas.get_width_height()
    img_array = np.fromstring(buf, dtype=np.uint8).reshape(nrows, ncols, 4)

    # save figure size
    figsize = fig.get_size_inches()
    # we want 700wx400h
    fig.set_size_inches(( 700/fig.dpi, 400/fig.dpi ))
    
    CCstr = Util_WrapText("Pearson Correlation Coefficient: " + str(data.CorCoeff), 550, FONT, FONTSIZE)
    pvalstr = Util_WrapText("P-Value: " + str(data.pval), 550, FONT, FONTSIZE)
    slopestr = Util_WrapText("Slope: " + str(data.slope), 550, FONT, FONTSIZE)
    interceptstr = Util_WrapText("Intercept: " + str(data.intercept), 550, FONT, FONTSIZE)

    # ugly ugly ugly way to copy over the figure, but any other method I tried yeilded bit jobbled garbage
    save_image = Image.new("RGBA", (700, 400+500) , "#ffffffff")
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi = fig.dpi)
    buf.seek(0)
    mplimage = deepcopy(Image.open(buf))
    buf.close()
    
    save_image.paste(mplimage, (0,0))

    drawtool = ImageDraw.Draw(save_image) # paste figure onto new image that we will be saving

    font = ImageFont.truetype(FONT, size=FONTSIZE)

    heightoffset = 30
    left, top, right, bottom = drawtool.textbbox((0,0), CCstr, font=font)
    drawtool.text((20, 400 + heightoffset), CCstr, fill=(0,0,0,255))
    heightoffset = heightoffset + bottom+10
    
    left, top, right, bottom = drawtool.textbbox((0,0), pvalstr, font=font)
    drawtool.text((20, 400 + heightoffset), pvalstr, fill=(0,0,0,255))
    heightoffset = heightoffset+ bottom+10
    
    left, top, right, bottom = drawtool.textbbox((0,0), slopestr, font=font)
    drawtool.text((20, 400 + heightoffset), slopestr, fill=(0,0,0,255))
    heightoffset = heightoffset + bottom+10
    
    left, top, right, bottom = drawtool.textbbox((0,0), interceptstr, font=font)
    drawtool.text((20, 400 + heightoffset), interceptstr, fill=(0,0,0,255))
    heightoffset = heightoffset + bottom+10
    
    for key,value in data.metadata.items():
        tempstr = Util_WrapText(str(key) + " : " + str(value), 550, FONT, FONTSIZE)

        left, top, right, bottom = drawtool.textbbox((0,0), tempstr, font=font)
        drawtool.text((20, 400 + heightoffset), tempstr, fill=(0,0,0,255))
        heightoffset = heightoffset + bottom+10

    save_image.save(filepath)
    
    fig.set_size_inches(figsize) # set size back to what we saved

    
def _QuitDataViewer(root):
    root.quit()
    root.destroy()

# contatiner for data to implement sort functions and visualization
# Create different modes of data visualization, such as viewing multiple regressions at the same time
class ExpressionDataContainer:
    def __init__(self):
        self.Data = []
        self.tkwindow = None
        self.metadata = []

    def __init__(self, data):
        self.Data = data
        self.metadata = []
        
    def ImportData(self, data):
        self.Data = data
        self.metadata = data.metadata

    # NOTE: processed data buffer NEEDS to be copied for new data container so we don't accidentally
    # mutate the raw data that we very much want to keep constant so it isn't accidentally
    # saved to a file and the original values arent overwritten
    @staticmethod
    def CreateForClinicalFeature(feature,val):
        return ExpressionDataContainer(filter(lambda x: x.GetMetadata(feature) == val, ProcessedDataBuffer.copy()))

    @staticmethod
    def CreateForGeneList(Igenes, Dgenes):
        return ExpressionDatacontainer(filter(lambda x: x.IGene in Igenes or x.DGene in Dgenes, ProcessedDataBuffer.copy()))
    
    @staticmethod
    def CreateFromDefault():
        return ExpressionDataContainer(ProcessedDataBuffer.copy())
        
    # returns container of sorted data from lowest to highest, opposite if inv is true
    def sort(self, Fn, inv=False):
        if not inv:
            self.Data.sort(key=Fn)
        else:
            self.Data.sort(key=Fn)
            self.Data.reverse()

    def copy(self):
        copy = ExpressionData()
        temp = self.Data.copy()
        copy.ImportData(temp)
        return copy

    def filter(self, Fn):
        # returns a copy
        return ExpressionDataContainer(list(filter(Fn, self.Data)))

    @staticmethod
    def _FilterSmallData(data):
        # filter out data with less than 10 values
        if len(data.IValues) < 30:
            return False
        return True

    def FilterSmallData(self):
        return self.filter(ExpressionDataContainer._FilterSmallData)

    @staticmethod
    def _FilterNSPVal(data):
        #filter out data with a pvalue greater than .05
        if data.pval > .05:
            return False
        return True

    def FilterNSPVal(self):
        return self.filter(ExpressionDataContainer._FilterNSPVal)
    
    @staticmethod
    def _FilterLowCC(data):
        #filter out data with a correlation coefficient -.3 < x <.3
        if abs(data.CorCoeff) < .3:
            return False

        return True

    def FilterLowCC(self):
        return self.filter(ExpressionDataContainer._FilterLowCC)

    def FilterForGenes(self, genes):
        return self.filter(lambda x: x.IGene in genes or x.DGene in genes)
    
    def SearchForGene(self, gene):
        returnVal = []
        for val in self.Data:
            if val.DGene == gene:
                returnVal.append(val)

        return ExpressionDataContainer(returnVal)


    def FilterForMetadata(self, ID, val):
        return self.filter(lambda x: x.metadata[ID] == val)

    #def GetMeanOfMetric(metricfn):

    #def GetMedianOfMetric(metricfn):

    #def GetSDOfMetric(metricfn):

    # assume predarr is sorted by which most likely targets are first, and least likely are last
    def MatchMiRTargetPredData(self, name, predarr):
        for val in self.Data:
            if val.DGene in predarr:
                val.metadata[MIR_PREDICTION_ALGORITHM].append(name)
                val.metadata[str(name + " priority")] = predarr.index(val.DGene)


    def FilterByPredAlgorithm(self):
        return self.filter(lambda x: len(x.metadata[MIR_PREDICTION_ALGORITHM]) > 0)

    def _FilterAndSaveItems(self, child_win, selection):
        directory = filedialog.askdirectory(title="Select Directory", mustexist=True)
        print(directory)
        
        # temp figure for subplots
        fig,ax = plt.subplots()
        
        for val in selection.curselection():
            # get index, which are the characters after the " "
            graphname = selection.get(val)
            str_idx = graphname.rfind(" ")+1

            data_idx = int(graphname[str_idx:])
            graphdata = self.Data[data_idx]

            print(data_idx)
            
            # create file path
            if directory[-1] != '/':
                directory += '/'

            # remove tricky characters and build filename
            temptrans = str.maketrans("", "", "?!.")
            
            filepath = directory + graphname[0:str_idx-1].translate(temptrans) + ".png"
            
            graphdata.PlotDataForEmbedded(fig, ax)

            print(filepath)
            
            _SaveWindow(child_win, fig, graphdata, filepath)
    

    def _SaveVisualDialog(self):
        if self.tkwindow is None:
            return
        
        # create child window
        child_win = Toplevel(self.tkwindow)
        child_win.title("Save Dialog")
        child_win.geometry("600x300")
        
        # create check buttons for each data
        data_strs = [data.IGene + " vs. " + data.DGene + " " + str(index) for index,data in enumerate(self.Data)]
        
        # create scrollbar
        scrollbar = Scrollbar(child_win, orient='vertical')
        scrollbar.pack(side=RIGHT, fill=Y)
        
        listbox = Listbox(child_win, selectmode='multiple', yscrollcommand = scrollbar.set)
        listbox.pack(padx=10,pady=10,expand=YES,fill="both")

        for data_str in data_strs:
            listbox.insert(END, data_str)
        
        
        SaveButton = Button(child_win, text="Save", command=lambda: self._FilterAndSaveItems(child_win, listbox))
        #CancelButton = Button(child_win, text="Cancel", command=lambda: child_win.destroy())
        
        SaveButton.pack(side=BOTTOM, fill=X)
        #CancelButton.pack(side=BOTTOM, fill=X)
        
        child_win.protocol("WM_DELETE_WINDOW", lambda: child_win.destroy())
        
        # display everything
        child_win.mainloop()
        

    def VisualizeData(self):
        # click through each data set, which displays a scatterplot with a linear regression, and any other metadata
        root_win = Tk()
        root_win.title("Expression Data")

        self.tkwindow = root_win
        #root_win.geometry("700x1600")
        
        '''becuase there is no explicit referencing/copying in python all variables that are changed
            by event function must be in an array or dictionary so they can be changed... >:('''
        data = [0] # index of expression data, list box data structure, figure, ax, toolbar

        #first embed matplotlib window
        fig, ax = self.Data[0].GetSubplots()
        fig.suptitle(str(self.Data[0].IGene) + " vs. " + str(self.Data[0].DGene))
        canvas = FigureCanvasTkAgg(fig, master=root_win)  # A tk.DrawingArea.
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        toolbar = NavigationToolbar2Tk(canvas, root_win)
        toolbar.update()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        scrollbar = Scrollbar(root_win)
        scrollbar.pack( side = RIGHT )

        valuelist = Listbox(root_win, yscrollcommand = scrollbar.set, width=100 )
        valuelist.insert(END, "Correlation Coefficient: " + str(self.Data[0].CorCoeff))
        valuelist.insert(END, "p-value: " + str(self.Data[0].pval))
        valuelist.insert(END, "Regression line slope: " + str(self.Data[0].slope))
        valuelist.insert(END, "Regression line intercept: " + str(self.Data[0].intercept))
        
        for key, value in self.Data[0].metadata.items():
           valuelist.insert(END, key + ": " + str(value))
        
        valuelist.insert(END, str(data[0]+1) + "/" + str(len(self.Data)))
        
        valuelist.pack( side = LEFT, fill = BOTH )
        scrollbar.config( command = valuelist.yview )

        data.append(valuelist)
        data.append(fig)
        data.append(ax)
        data.append(toolbar)

        prevbutton = Button(root_win, text="next", command=lambda: _ExpressionDataVisualizerNavigateRight(self, data, root_win))
        nextbutton = Button(root_win, text="prev", command=lambda: _ExpressionDataVisualizerNavigateLeft(self, data, root_win))
        savebutton = Button(root_win, text="save", command=lambda: self._SaveVisualDialog())
        
        prevbutton.pack(side=RIGHT)
        nextbutton.pack(side=RIGHT)
        savebutton.pack(side=RIGHT)
        
        root_win.protocol("WM_DELETE_WINDOW", lambda: _QuitDataViewer(root_win))
        
        tk.mainloop()

        self.tkwindow = None
        
def ProcessAllGenesForIGene(IGene, Iexpressiontype, Dexpressiontype):
    global ProcessedDataBuffer
    ProcessedDataBuffer.clear()
    with alive_bar(len(GeneList)) as bar:
        for val in GeneList:
            if IGene == val: continue
            tempgraph = ExpressionData(IGene, Iexpressiontype, val, Dexpressiontype)
            tempgraph.ProcessData()
            ProcessedDataBuffer.append(tempgraph)
            bar()

# returns an array of processed genes seperated by the category in the clinical feature
def ProcessGeneForClinicalFeature(IGene, Iexpressiontype, Dexpressiontype, feature):
    global ProcessedDataBuffer
    ProcessedDataBuffer.clear()
    # first iterate through patient data, gather possible clinical feature values
    featurestates = []
    for pt in TCGA_Patients:
        temp = pt.GetClinicalFeature(feature)
        if temp is None:
            continue

        if type(temp) is not str:
            continue #sometimes pandas will be dumb and replace NA with NAN even when I want NA to mean 'not applicable'
        
        if temp not in featurestates:
            featurestates.append(temp)
    
    # for each possible value, compute expression data of that group
    with alive_bar(len(GeneList) * len(featurestates)) as bar:
        for featureval in featurestates:
            for val in GeneList:
                if IGene is val: continue
                tempgraph = ExpressionData(IGene, Iexpressiontype, val, Dexpressiontype)
                tempgraph.ProcessDataForClinicalFeature(feature, featureval)
                ProcessedDataBuffer.append(tempgraph)
                bar()
    
    
''' TODO: CITE THESE DATABASES '''

# one might want to use some prediction algorithms to narrow the search for miRNA's of interest, so thats what were gonna do
#_PicTarRequestURL = 'https://pictar.mdc-berlin.de/cgi-bin/PicTar_vertebrate.cgi'
'''def _PicTarSubmitRequest(miRNA):
    #TODO: gotta captialize the r in miR or else the request won't recognize it
    
    
    requestpayload = {'clade' : 'vertebrate',
                      'dataset' : 'target predictions for all human microRNAs based on conservation in mammals (human, chimp, mouse, rat, dog)',
                      'name2' : miRNA,
                      'name1' : '',
                      'action' : 'Search for targets of a miRNA'}
    
    r = requests.post(_PicTarRequestURL, data=requestpayload)

    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")
        
    return r.text
    '''
# returns a list of the predicted genes, REQUIRES GeneDict AND GeneList TO BE LOADED INTO MEMORY
def _PicTarGetPredictionSet(miRNA):
    requestpayload = {'clade' : 'vertebrate',
                      'dataset' : 'target predictions for all human microRNAs based on conservation in mammals (human, chimp, mouse, rat, dog)',
                      'name2' : miRNA,
                      'name1' : '',
                      'action' : 'Search for targets of a miRNA'}
    
    r = requests.post(_PicTarRequestURL, data=requestpayload)

    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")
    
    soup = BeautifulSoup(r.text, 'html.parser')
    table = soup.find('table').find_next('table')
    RetList = []
    
    for val in table.contents[3:]:
        #iterate through each entry in the table
        genestr = val.contents[-4].string # this way of indexing is some vodoo stuff i just kept subtracting 1 until it gave me what I wanted lol
        
        openpos = genestr.find('(')
        closepos = genestr.find(')')
        if openpos == -1 or closepos == -1:
            continue
        
        while closepos < len(genestr):
            if openpos > closepos:
                closepos = genestr.find(')', closepos)
                if closepos == -1:
                    temp = None
                    break;

            temp = genestr[openpos+1 : closepos]

            if temp in GeneList:
                break
            
            if temp in GeneDict.keys():
                if GeneDict[temp] in GeneList:
                    temp = GeneDict[temp]
                    break
            
            openpos = genestr.find('(', openpos+1)
            closepos = genestr.find(')', closepos+1)

            if openpos == -1 or closepos == -1:
                temp = None
                break
        
        if temp is None: continue
        
        # the gene is usually in parenthases, so use that as a way to tell for now
        RetList.append(temp)
        
    return RetList

# structure that defines standard group of operations that can be done on a webpage
@dataclass
class PredictionSet:
    def __init__(self, name, predfunc):
        self.name = name
        self.predfunc = predfunc
        
_MirDBRequestURL = 'https://mirdb.org/cgi-bin/search.cgi'
_MirDBBaseURL = 'https://mirdb.org'
'''def _MirDBSubmitRequest(miRNA):
    requestpayload = {'species' : 'Human',
                      'searchBox' : miRNA,
                      'submitButton' : 'Go',
                      'searchType' : 'miRNA'}
    
    r = requests.post(_MirDBRequestURL, data=requestpayload)

    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")

    return r.text
'''
def cleanhtml(text):
  cleantext = re.sub(re.compile('<.*?>'), '', text)
  return cleantext

def _MirDBParseWindowFunc(selection, combobox, root):
    selection.append(combobox.get())
    root.quit()

def _MirDBGetPredictionSet(miRNA):
    requestpayload = {'species' : 'Human',
                      'searchBox' : miRNA,
                      'submitButton' : 'Go',
                      'searchType' : 'miRNA'}
    
    r = requests.post(_MirDBRequestURL, data=requestpayload)

    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")

    soup = BeautifulSoup(r.text, 'html.parser')
    # first check if it directed us to results or another page letting us specify by checking if there is an h3
    if soup.find('h3'):
       # yep, multiple entries, make a window popup and ask the user which one they want
       
        temp = soup.find(href=re.compile('searchType=miRNA'))
        mirlist = temp.find_all('h4')
        for i in range(len(mirlist)):
            mirlist[i] = cleanhtml(mirlist[i].string)
        
        selection = [] # have to use an array because python doesn't have explicit references :(
        root_win = Tk()

        label = Label(root_win, text='Multiple genes have been returned from search, please select desired gene')
        label.pack()
        combobox = Combobox(root_win, state="readonly", values=mirlist)
        combobox.pack()
        button = Button(root_win, text='OK', command=lambda: _MirDBParseWindowFunc(selection, combobox, root_win) )
        button.pack()
        
        root_win.mainloop()

        root_win.destroy()
        # now make the request using the selection
        urltag = temp.find(string=selection[0]).parent.parent
        r = requests.get(_MirDBBaseURL + urltag['href'])
        if not r.status_code == 200:
            warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")

        soup = BeautifulSoup(r.text, 'html.parser')

    table = soup.find('table', id='table1')
    RetList = []
    
    for val in table.contents[2:]:
        #iterate through each entry in the table
        # every other value is '\n', but its not a big deal
        if val.string == '\n':
            continue
        
        if not val.find('a', target='_blank'):
            continue
        
        temp = val.find('a', target='_blank').string
        temp = temp.strip()

        if temp in GeneList:
            RetList.append(temp)

        if temp in GeneDict.keys():
            if GeneDict[temp] in GeneList:        
                RetList.append(GeneDict[temp])
        
    return RetList

_TargetScanRequestURL = 'https://www.targetscan.org/cgi-bin/targetscan/vert_80/targetscan.cgi?'
_TargetScanBaseURL = 'https://www.targetscan.org'
'''def _TargetScanSubmitRequest(miRNA):
    payload = {'species' : 'Human',
               'gid' : '',
               'mir_sc' : '',
               'mir_nc' : '',
               'mir_vnc' : '',
               'mirg' : miRNA}

    # GET request >:(
    r = requests.get(_TargetScanRequestURL + urllib.parse.urlencode(payload))
    
    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")

    return r.text
'''
def _TargetScanGetPredictionSet(miRNA):
    payload = {'species' : 'Human',
               'gid' : '',
               'mir_sc' : '',
               'mir_nc' : '',
               'mir_vnc' : '',
               'mirg' : miRNA}

    # GET request >:(
    r = requests.get(_TargetScanRequestURL + urllib.parse.urlencode(payload))
    
    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")
    
    soup = BeautifulSoup(r.text, 'html.parser')
    # first check if it directed us to results or another page letting us specify by checking if there is an h3
    if soup.find(string=re.compile('matches multiple families in our miRNA database')):
        # yep, multiple entries, make a window popup and ask the user which one they want
        templist = soup.find('h3').find('table').find_all('a')
        mirlist = []
        for val in templist:
            mirlist.append(val.string)
        
        selection = [] # have to use an array because python doesn't have explicit references :(
        root_win = Tk()

        label = Label(root_win, text='Multiple genes have been returned from search, please select desired gene')
        label.pack()
        combobox = Combobox(root_win, state="readonly", values=mirlist)
        combobox.pack()
        button = Button(root_win, text='OK', command=lambda: _MirDBParseWindowFunc(selection, combobox, root_win) )
        button.pack()
        
        root_win.mainloop()

        root_win.destroy()
        # now make the request using the selection
        urltag = soup.find('h3').find('a',string=selection[0])
        
        r = requests.get(_TargetScanBaseURL + urltag['href'])
        if not r.status_code == 200:
            warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")
        soup = BeautifulSoup(r.text, 'html.parser')

    genetags = soup.find('table', id='restable').find_all('a',target='new')
    RetList = []

    for val in genetags:
        #iterate through each entry in the table
        # every other value is '\n', but its not a big deal
        if val.string == '\n':
            continue

        temp = val.string

        if temp in GeneList:
            RetList.append(temp)

        if temp in GeneDict.keys():
            if GeneDict[temp] in GeneList:
                RetList.append(GeneDict[temp])
        
    return RetList

_MirWalkRequestURL = 'http://mirwalk.umm.uni-heidelberg.de/search/?'
'''def _MirWalkSubmitRequest(miRNA):
    payload = {'species' : 'human', 'gene' : '', 'mirna' : miRNA}

    r = requests.get(_MirWalkRequestURL + urllib.parse.urlencode(payload))

    if not r.status_code == 200:
        warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")

    return r.text
'''
def _MirWalkGetPredictionSet(miRNA):
    with request.Session() as session:
        payload = {'species' : 'human', 'gene' : '', 'mirna' : miRNA}
        r = requests.get(_MirWalkRequestURL + urllib.parse.urlencode(payload))

        if not r.status_code == 200:
            warnings.warn("Status code of the returned data is not OK! Data may be incorrect or corrupt!")

        # create session up here
        soup = BeautifulSoup(r.text, 'html.parser')

        if soup.find(string=re.compile('was not found in our database')):
            templist = soup.find(class_='table-container').find_all('tr')
            mirlist = []
            saveurl = {}
            for val in templist[1:]:
                temp = val.find_all('a')[1]
                mirlist.append(temp.string)
                saveurl[mirlist[-1]] = temp['href']
                
            selection = [] # have to use an array because python doesn't have explicit references :(
            root_win = Tk()
                
            label = Label(root_win, text='Multiple genes have been returned from search, please select desired gene')
            label.pack()
            combobox = Combobox(root_win, state="readonly", values=mirlist)
            combobox.pack()
            button = Button(root_win, text='OK', command=lambda: _MirDBParseWindowFunc(selection, combobox, root_win) )
            button.pack()
                
            root_win.mainloop()
                
            root_win.destroy()
            # now make the request using the selection
            
        r = session.get(saveurl[selection[0]])
        data = b''
        with session.get('http://mirwalk.umm.uni-heidelberg.de/export/csv/', stream=True) as response:
            with alive_bar(int(response.headers.get('Content-Length'))/4096) as bar:
                for chunk in response.iter_content(chunk_size=4096):
                    if chunk:
                        data = data + chunk
                        bar()
        
        
        df = pd.read_csv(StringIO(data.encode('UTF-8')))

        df = df[ df['bindingp'] > 0.95]
        
        return df['mirnaid']

class MirandaGeneData:
    def __init__(self):
        self.MirName = ""
        self.TargetName = ""
        self.TotalScore = ""
        self.TotalEnergy = ""

def ReadMirandaOutFile():
    file = filedialog.askopenfile(mode="r")

    if file is None:
        return None
    
    data = file.read().replace('\n', '')
    returnVal = []
    
    barpos=0
    fstart = 0
    fend = len(data)-1
    with alive_bar(fend) as bar:
        while True:
            mirdata = MirandaGeneData()
            
            index = data.find('>>', fstart, fend)
            if index == -1:
                while barpos<fend:
                    bar()
                    barpos = barpos+1
                    
                break
            
            index = index+2 # set position to start of next name

            #temp = data.find('\t', index, fend)
            mirdata.MirName = data[index:temp]

            index = temp+1
            temp = data.find('\t', index, fend)
            mirdata.TargetName = data[index:temp]

            index = temp+1
            temp = data.find('\t', index, fend)
            mirdata.TotalScore = data[index:temp]

            index = temp+1
            temp = data.find('\t', index, fend)
            mirdata.TotalEnergy = data[index:temp]

            returnVal.append(mirdata)

            fstart = temp
            
            while barpos<fstart:
                bar()
                barpos = barpos+1
    
    # sort based on score energy
    returnVal.sort(key=lambda x : x.TotalScore, reverse=True)
    return returnVal

RefseqConversionTable = None
def LoadRefseqConversionTable():
    global RefseqConversionTable
    
    file = filedialog.askopenfilename()
    RefseqConversionTable = pd.read_csv(file)

def GetGeneIDFromRefseq(refseq):
    if RefseqConversionTable is None:
        print("Please load conversion table")
        return None
    try:
        return RefseqConversionTable.loc[RefseqConversionTable['name'] == refseq]['name2'].values[0]
    except:
        print("it seems that either the refseq ID is either not in the table, or the table is not formatted correctly")
        return None

def _MirandaGetPredictionSet():
    mirandadata = ReadMirandaOutFile()
    strdata = []
    
    if mirandadata is None:
        return None

    if RefseqConversionTable is None:
        print("please load refseq conversion table")
        return None

    for element in mirandadata:
        element.TargetName = GetGeneIDFromRefseq(element.TargetName)
        strdata.append(element.TargetName)

    return mirandadata, strdata

genecopoia_URL = "https://www.genecopoeia.com/product/search/result.php?pageNum_Recordset1=0&field=11&tax_id=10090&key=mmu-miR-196b-5p&prt=16&totalRows_Recordset1=252"
genecopoia_baseURL = "https://www.genecopoeia.com"

def _AuxGenecopiaFindPageIndex(pages, index):
    for page in pages:
        try:
            if (int(page.text) == index): return page
        except ValueError as ve:
            continue

# these targets are for mouse, no simple way to navigate site so just provide URL of target query results
def _GetGenecopiaMMTargets():
    if GeneDict == {}:
        print("The gene dictionary is not loaded!")
        return
    
    filename = filedialog.askopenfilename()
    cookies = None
    returnVal = []

    with open(filename, 'rb') as f:
        cookies = pickle.load(f)

    c_jar = requests.cookies.RequestsCookieJar()
    
    for cookie in cookies:
        c_jar.set(cookie['name'], cookie['value'], domain=cookie['domain'], path=cookie['path'])
    
    r = requests.get(genecopoia_URL,cookies=c_jar)

    soup = BeautifulSoup(r.text, 'html.parser')

    pages = soup.find(class_='pagination').find_all('li')
    
    # find total number of pages
    numpages = int(pages[-1].text)
    
    for i in range(2, numpages+1):
    # for each page (including starting page)
        
        #   get body of results
        table = soup.find(class_="table_org dcf-table dcf-table-responsive dcf-table-bordered dcf-table-striped dcf-w-100%").find_all('tr')
        pages = soup.find(class_='pagination').find_all('li')

        # skip first row
        for idx in range(1, len(table)):
            # get all RNA names given (data-label: "Symbol"), or position 3 in .contents
            gene = table[idx].contents[3].text.upper()
            
            if gene in GeneDict.keys():
                if GeneDict[gene] in GeneList:        
                    returnVal.append(GeneDict[gene])
        
            
        #   navigate to next page
        link = genecopoia_baseURL + _AuxGenecopiaFindPageIndex(pages, i).a['href']
        
        #   repeat
        soup = BeautifulSoup(requests.get(link, cookies=c_jar).text, 'html.parser')


    return returnVal

# keep in mind these values will be saved as strings
def LoadFirebrowseClinicalFeatures():
    path = filedialog.askopenfilename()
    
    # is the file excel or tsv?
    df = pd.read_csv(path, sep='\t', keep_default_na=False)
    # create dataframe from excel file (or tsv)
    global TCGA_Patients

    # for each patient (TCGA-##-####)
    #   append each feature to clinical feature list yes, no, N/A <-- frustrating :(
    for pt in TCGA_Patients:
        try:
            temprow = df[pt.ID.lower()]
        except:
            continue

        # set colum headers to their descriptors
        temprow.index = df['bcr_patient_barcode']

        for value in temprow.index:
            pt.AppendClinicalFeature(value, temprow[value])

def GetPossibleFeatureValues(clinical_feature):
    global TCGA_Patients
    clinical_feature_values = []

    
    for pt in TCGA_Patients:
        feature = pt.GetClinicalFeature(clinical_feature)
        if feature not in clinical_feature_values:
            clinical_feature_values.append(feature)

    return clinical_feature_values

def BuildFeatureList(clinical_feature):
    returnVal = {}

    for pt in TCGA_Patients:
        feature = pt.GetClinicalFeature(clinical_feature)
        if not feature:
            continue

        if feature not in returnVal.keys():
            returnVal[feature] = [pt.ID]
        else:
            returnVal[feature].append(pt.ID)

    return returnVal

def prime_factors(n):
    returnVal = []

    i : int = 2
    while i*i <= n:
        if n % i:
            i += 1
            continue
        else:
            n //= i
            returnVal.append(i)

    if n > 1:
        returnVal.append(n)
    
    return returnVal
    
# only works well for small values, for values >= 15 it will give a very uneven split!
def GetClosestSquareBounds(n):
    factors = prime_factors(n)
        
    low = hi = 1

    for x in factors[:int(np.ceil(len(factors)/2))]:
        low *= x

    for x in factors[int(np.ceil(len(factors)/2)):]:
        hi *= x

    return (low, hi)

def GetGeneStratID(gene):
    return f"{gene}_expression"

def GenerateClinicalFeatureForGeneStratification(gene, expressiontype, low, hi):
    global TCGA_Patients

    expression = np.array([float(val.GetExpressionValue(gene, expressiontype)) for val in TCGA_Patients if val is not None and val.GetExpressionValue(gene, expressiontype) is not None])

    val_high = np.percentile(expression, hi)
    val_low = np.percentile(expression, low)
    
    for pt in TCGA_Patients:
        exp = pt.GetExpressionValue(gene, expressiontype)
        if exp is not None:
            if float(exp) > val_high:
                pt.AppendClinicalFeature(GetGeneStratID(gene), "hi")
            elif float(exp) < val_low:
                pt.AppendClinicalFeature(GetGeneStratID(gene), "lo")


def RemoveGeneStratificationFeature(gene):
    for pt in TCGA_Patients:
        pt.RemoveClinicalFeature(GetGeneStratID(gene))

def PlotExpressionByFeature(Igene, expressiontype, nullgroup, stratification_metric, plot : bool = True):
    cvalues = [x for x in GetPossibleFeatureValues(stratification_metric) if x]
    #pdb.set_trace()
    cvalmap = {x : idx for idx,x in enumerate(cvalues)}
    values = [[] for x in cvalues]
        
    for pt in TCGA_Patients:
        featureval = pt.GetClinicalFeature(stratification_metric)

        if featureval is None:
            continue
        if pt.CheckGeneDict(Igene, expressiontype):
            exp = float(pt.GetExpressionValue(Igene, expressiontype))
            if exp != 0.0:
                values[ cvalmap[ pt.GetClinicalFeature(stratification_metric) ] ].append( float(pt.GetExpressionValue(Igene, expressiontype)) )
       
    # check which groups have zero values and remove them ( don't combine, leave it like this for the sake of debugging and readability)
    usemetrics = [i for i in range(0,len(cvalmap)) if len(values[i]) > 0]

    if not usemetrics:
        print("No patients with clinical feature found!")
        return None
    
    if cvalmap[nullgroup] not in usemetrics:
        print("The null group is empty!")
        return None
    
    #avg = [ sum(values[idx]) / len(values[idx]) for idx in usemetrics ] # averages
    #nzf_rdpm = deepcopy(avg)
    # normalize each average (0 ulcerated, 1 non ulcerated)
    # iterate through usemetrics to only get groups that we KNOW have values
    for idx,pos in enumerate(usemetrics):
        for idx,val in enumerate(values[pos]):
            values[pos][idx] = np.log2(val)

    avg = [ sum(values[idx]) / len(values[idx]) for idx in usemetrics ] # averages
    
    
    err = [scipy.stats.sem(values[x]) for x in usemetrics]

    # go ahead and just return the significance if we are not plotting
    if plot is False:
        ttestafter = {}
        enriched = {}
        for idx,val in enumerate(usemetrics):
            if values[cvalmap[nullgroup]] == val:
                continue
            
            ttestafter[cvalues[val]] = scipy.stats.ttest_ind(values[cvalmap[nullgroup]], values[val])
            enriched[cvalues[val]] = avg[cvalmap[nullgroup]] > avg[idx]
        return (ttestafter, enriched)

    axesarea = len(usemetrics)
    bounds = GetClosestSquareBounds(axesarea)
    while 1 in bounds:
        axesarea += 1
        bounds = GetClosestSquareBounds(axesarea)
        
    #fig,ax = plt.subplots(ncols = bounds[0], nrows = bounds[1])
    fig,ax = plt.subplots()
    #plt.title(f'{Igene} expression in ulcerated vs non ulcerated melanoma seggregated by {stratification_metric}')
    plt.xlabel(stratification_metric)
    plt.ylabel('normalized RPKM')
    plt.title(f'{gene} expression stratified by {feature}')
    # iterate through each graph
    '''for idx,pos in enumerate(usemetrics): # only use the indicies that we know have values
        ax[idx // bounds[0], idx % bounds[0]].violinplot([values[pos], values[pos]])
        ax[idx // bounds[0], idx % bounds[0]].set_xticks([1,2], labels=cvalues)
            
        #for i in range(0,len(ulceration_stats)):
        #    ax[idx // bounds[0], idx % bounds[0]].scatter(i + np.random.random(len(values[i][pos])) * .8 - .8 / 2, values[i][pos], color=colors[i], label=ulceration_stats[i])


        ax[idx // bounds[0], idx % bounds[0]].set_title(f'{cvalues[pos]}')
        ax[idx // bounds[0], idx % bounds[0]].set_xlabel(stratification_factor)
        ax[idx // bounds[0], idx % bounds[0]].set_ylabel(f'log2 {Igene} Expression')'''

    ax.violinplot([val for val in values if val])
    ax.set_xticks(range(1,len(cvalues)+1), labels=cvalues)
    
    ttestafter = {}
    enriched = {}

    for idx,val in enumerate(usemetrics):
        if values[cvalmap[nullgroup]] == val:
            continue
        
        ttestafter[cvalues[val]] = scipy.stats.ttest_ind(values[cvalmap[nullgroup]], values[val])
        enriched[cvalues[val]] = avg[cvalmap[nullgroup]] < avg[idx]

    print(ttestafter)
        
    plt.subplots_adjust(hspace=1.5)
    plt.show()

    return (ttestafter, enriched)

def UlcerationGraph(Igene, expressiontype, stratification_metric=None, plot : bool = True):
    
    #pdb.set_trace()
    
    ulceration_stats = ['Ulcerated', 'Non-ulcerated']
    '''stranslate = {'Local Disease (i & ii)' : ['stage 0','stage i','stage ia','stage ib','stage ic', 'i/ii nos', 'stage ii', 'stage iia', 'stage iib', 'stage iic'],
                  'Late Stage and Mestastatic (iii & iv)' : ['stage iii', 'stage iiia', 'stage iiib', 'stage iiic', 'stage iv'],
                  'Only Metastatic (iv)' : ['stage iv'],
                  'All Stages' : stages}'''

    colors = [(0, 0, 1, 1), (1, 0, 0, 1)]
    if stratification_metric is None:
        values = [ [],[] ]
        for pt in TCGA_Patients:
            if pt.ClinicalFeatures['melanoma_ulceration_indicator'] == 'yes':
                if pt.CheckGeneDict(Igene, expressiontype):
                    exp = float(pt.GetExpressionValue(Igene, expressiontype))
                     
                    if exp != 0.0:
                        values[0].append( float(pt.GetExpressionValue(Igene, expressiontype)) )
            elif pt.ClinicalFeatures['melanoma_ulceration_indicator'] == 'no':
                if pt.CheckGeneDict(Igene, expressiontype):
                    exp = float(pt.GetExpressionValue(Igene, expressiontype))
                     
                    if exp != 0.0:
                        values[1].append( float(pt.GetExpressionValue(Igene, expressiontype)) )

        #err = [scipy.stats.sem(values[0]), scipy.stats.sem(values[1])]

        counts = [sum(values[0]) / len(values[0]), sum(values[1]) / len(values[1])]
        #print(counts)
        # normalize to fold change
        nzf_rdpm = counts[1]

        #ttestbefore = scipy.stats.ttest_ind(values[0], values[1])
        
        for index,val in enumerate(values[0]):
            values[0][index] = float(np.log2(val))#np.log2(val/nzf_rdpm)

        for index,val in enumerate(values[1]):
            values[1][index] = float(np.log2(val))#np.log2(val/nzf_rdpm)


        if plot is False:
            ttestafter = scipy.stats.ttest_ind(values[0], values[1])
            return (ttestafter, counts[0] > counts[1])
        

        fig,ax = plt.subplots()
        
        # recompute mean
        counts = [sum(values[0]) / len(values[0]), sum(values[1]) / len(values[1])]
        
        # calculate standard error of mean
        err = [scipy.stats.sem(values[0]), scipy.stats.sem(values[1])]
        ax.violinplot(values)
        
        #plt.errorbar(ulceration_stats, [counts[0], counts[1]], yerr=err, fmt="o", color="g")
        ax.set_xticks([1,2], labels=ulceration_stats)
        ax.set_ylabel(f'log2 of {expressiontype}')
        ax.set_title(f'log2 {Igene} expression in ulcerated vs non ulcerated Melanoma')

        #for i in range(len(ulceration_stats)):
             #ax.scatter(i + np.random.random(len(values[i])) * .8 - .8 / 2, values[i], color=colors[i])

        ttestafter = scipy.stats.ttest_ind(values[0], values[1])
        print(ttestafter)
        plt.show()
        
        return (ttestafter, counts[0] > counts[1])
    else:
        cvalues = GetPossibleFeatureValues(stratification_metric)
        #pdb.set_trace()
        cvalmap = {x : idx for idx,x in enumerate(cvalues)}
        values = [[[] for x in cvalues], [[] for x in cvalues]]
        
        for pt in TCGA_Patients:
            if pt.ClinicalFeatures['melanoma_ulceration_indicator'] == 'yes': #TODO: This should be .GetClinicalFeature
                if pt.CheckGeneDict(Igene, expressiontype):
                    if pt.GetClinicalFeature(stratification_metric) is None:
                        continue
                    
                    exp = float(pt.GetExpressionValue(Igene, expressiontype))
                    if exp != 0.0:
                        values[0][ cvalmap[ pt.GetClinicalFeature(stratification_metric) ] ].append( float(pt.GetExpressionValue(Igene, expressiontype)) )
            elif pt.GetClinicalFeature('melanoma_ulceration_indicator') == 'no':
                if pt.CheckGeneDict(Igene, expressiontype):
                    if pt.GetClinicalFeature(stratification_metric) is None:
                        continue
                    
                    exp = float(pt.GetExpressionValue(Igene, expressiontype))
                    if exp != 0.0:
                        values[1][ cvalmap[ pt.GetClinicalFeature(stratification_metric) ] ].append( float(pt.GetExpressionValue(Igene, expressiontype)) )
                    
        # check which groups have zero values and remove them ( don't combine, leave it like this for the sake of debugging and readability)
        usemetrics = [i for i in range(0,len(cvalmap)) if len(values[0][i]) > 0 and len(values[1][i]) > 0]
        
        avg = [ [ sum(values[0][idx]) / len(values[0][idx]) for idx in usemetrics ], [ sum(values[1][idx]) / len(values[1][idx]) for idx in usemetrics ] ] # averages
        nzf_rdpm = deepcopy(avg[1]) # non-uncerated groups
        # normalize each average (0 ulcerated, 1 non ulcerated)
        # iterate through usemetrics to only get groups that we KNOW have values
        for idx0,pos0 in enumerate(usemetrics):
            #avg[0][idx0] = avg[0][idx0] / nzf_rdpm[idx0]
            for idx,val in enumerate(values[0][pos0]):
                values[0][pos0][idx] = np.log2(val)# / nzf_rdpm[idx0])

        for idx1,pos1 in enumerate(usemetrics):
            #avg[1][idx1] = 1.0
            for idx,val in enumerate(values[1][pos1]):
                values[1][pos1][idx] = np.log2(val )#/ nzf_rdpm[idx1])

        # recompute averages
        avg = [ [ sum(values[0][idx]) / len(values[0][idx]) for idx in usemetrics ], [ sum(values[1][idx]) / len(values[1][idx]) for idx in usemetrics ] ]
    
        
        err = [[scipy.stats.sem(values[0][x]) for x in usemetrics], [scipy.stats.sem(values[1][x]) for x in usemetrics]]

        # go ahead and just return the significance if we are not plotting
        if plot is False:
            ttestafter = {}
            enriched = {}
            for idx,val in enumerate(usemetrics):
                ttestafter[cvalues[val]] = scipy.stats.ttest_ind(values[0][val], values[1][val])
                enriched[cvalues[val]] = avg[0][idx] > avg[1][idx]
            return (ttestafter, enriched)

        axesarea = len(usemetrics)
        bounds = GetClosestSquareBounds(axesarea)
        while 1 in bounds:
            axesarea += 1
            bounds = GetClosestSquareBounds(axesarea)
        
        fig,ax = plt.subplots(ncols = bounds[0], nrows = bounds[1])
        #plt.title(f'{Igene} expression in ulcerated vs non ulcerated melanoma seggregated by {stratification_metric}')
        plt.xlabel('Ulceration Status')
        plt.ylabel('normalized RPKM')
        # iterate through each graph
        for idx,pos in enumerate(usemetrics): # only use the indicies that we know have values
            ax[idx // bounds[0], idx % bounds[0]].violinplot([values[0][pos], values[1][pos]])
            ax[idx // bounds[0], idx % bounds[0]].set_xticks([1,2], labels=ulceration_stats)
            
            #for i in range(0,len(ulceration_stats)):
            #    ax[idx // bounds[0], idx % bounds[0]].scatter(i + np.random.random(len(values[i][pos])) * .8 - .8 / 2, values[i][pos], color=colors[i], label=ulceration_stats[i])


            ax[idx // bounds[0], idx % bounds[0]].set_title(f'{cvalues[pos]}')
            ax[idx // bounds[0], idx % bounds[0]].set_xlabel(f'{stratification_metric}: {cvalues[pos]}')
            ax[idx // bounds[0], idx % bounds[0]].set_ylabel(f'log2 {Igene} Expression')

        ttestafter = {}
        enriched = {}

        for idx,val in enumerate(usemetrics):
            ttestafter[cvalues[val]] = scipy.stats.ttest_ind(values[0][val], values[1][val])
            enriched[cvalues[val]] = avg[0][idx] > avg[1][idx]

        print(ttestafter)
        
        plt.subplots_adjust(hspace=1.5)
        plt.show()

        return (ttestafter, enriched)

def GuessExpressionType(gene):
    if '|' in gene:
        return 'normalized_count'
    else:
        return 'reads_per_million_miRNA_mapped'

# either hi or low, filtering genes that have higher or lower expression in ulcerated groups respectively
def FindSignificantUlcerationGenes(pval, df):
    global GeneList

    returnVal = []

    with alive_bar(len(GeneList)) as bar:
        for gene in GeneList:
            try:
                ttest, enriched = UlcerationGraph(gene, GuessExpressionType(gene), plot = False)
                #print(ttest)
                if df == 'hi':
                    if ttest.pvalue < pval and enriched:
                        returnVal.append((gene, ttest.pvalue))
                    bar()
                elif df == 'low':
                    if ttest.pvalue < pval and not enriched:
                        returnVal.append((gene, ttest.pvalue))
                    bar()
                elif df == 'both':
                    if ttest.pvalue < pval:
                        returnVal.append((gene, ttest.pvalue))
                    bar()
                else:
                    print(f'Invalid vaulue {df} passed to df, must be either hi, low or both!')
            except:
                bar()
    return returnVal

# find genes with differential expression between ulcerated and non ulcerated when miR expression is high
def FindDiffUlcerationGenes(pval, stratgene, lo, hi):
    global GeneList

    returnVal = []
    
    GenerateClinicalFeatureForGeneStratification(stratgene, GuessExpressionType(stratgene), lo, hi)
    
    with alive_bar(len(GeneList)) as bar:
        for gene in GeneList:
            try:
                ttest, enriched = UlcerationGraph(gene, GuessExpressionType(gene), stratification_metric=GetGeneStratID(stratgene),plot = False)
                if ttest['hi'].pvalue < pval and not enriched['hi']:
                    returnVal.append((gene, pval))
                bar()
            except:
                bar()
    #RemoveGeneStratificationFeature(stratgene)
    return returnVal

def FindDiffByMutation(pval, feature, nullgrp, df):
    global GeneList

    returnVal = []

    with alive_bar(len(GeneList)) as bar:
        for gene in GeneList:
            try:
                ttest, enriched = PlotExpressionByFeature(gene, GuessExpressionType(gene), nullgrp, feature, plot = False)
                #print(ttest)
                if df == 'hi':
                    if ttest['yes'].pvalue < pval and enriched['yes']:
                        returnVal.append((gene, ttest['yes'].pvalue))
                    bar()
                elif df == 'low':
                    if ttest['yes'].pvalue < pval and not enriched['yes']:
                        returnVal.append((gene, ttest['yes'].pvalue))
                    bar()
                elif df == 'both':
                    if ttest['yes'].pvalue < pval:
                        returnVal.append((gene, ttest['yes'].pvalue))
                    bar()
                else:
                    print(f'Invalid vaulue {df} passed to df, must be either \'hi\', \'low\' or \'both\'!')
            except:
                bar()
    return returnVal
    
# batch normalization
def CreateUlcerationWorkbook(mirs, readtypes):
    savename = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=(("excel file", "*.xlsx"),("All Files", "*.*") ))
    
    #create excel workbook
    wb = openpyxl.Workbook()
    # create ulcerated and non ulcerated sheets
    ulceratedws = wb.create_sheet('Ulcerated')
    nonulceratedws = wb.create_sheet('Non-Ulcerated')
    
    # this is kinda backwards... but it should work
    # interate through row of each column for each patient and segregate data
    for index,row in enumerate(ulceratedws.iter_rows(min_row=1, max_row=len(TCGA_Patients)+2, max_col=len(mirs) * len(readtypes)+1)):
        if index == 0:# title row 1, fill with gene names
            for gindex,gene in enumerate(mirs):
                row[gindex * len(readtypes) + 1].value = gene
            continue
        elif index == 1:
            for i in range(1,len(row)):
                row[i].value = readtypes[(i-1) % len(readtypes)]
            continue
        
        pt = TCGA_Patients[index-2]
        row[0].value = pt.ID # set index
        
        if pt.GetClinicalFeature('melanoma_ulceration_indicator') == 'yes':
            for gindex,gene in enumerate(mirs):
                for eindex,expression in enumerate(readtypes):
                    exprval = pt.GetExpressionValue(gene, expression)
                    if exprval is not None:
                        row[gindex*len(readtypes) + eindex + 1].value = str(np.log2(float(pt.GetExpressionValue(gene, expression))))
    for index,row in enumerate(nonulceratedws.iter_rows(min_row=1, max_row=len(TCGA_Patients)+2, max_col=len(mirs) * len(readtypes)+1)):
        if index == 0:# title row 1, fill with gene names
            for gindex,gene in enumerate(mirs):
                row[gindex * len(readtypes) + 1].value = gene
            continue
        elif index == 1:
            for i in range(1,len(row)):
                row[i].value = readtypes[(i-1) % len(readtypes)]
            continue
        
        pt = TCGA_Patients[index-2]
        row[0].value = pt.ID # set index
        
        if pt.GetClinicalFeature('melanoma_ulceration_indicator') == 'no':
            for gindex,gene in enumerate(mirs):
                for eindex,expression in enumerate(readtypes):
                    exprval = pt.GetExpressionValue(gene, expression)
                    if exprval is not None:
                        row[gindex*len(readtypes) + eindex + 1].value = str(np.log2(float(pt.GetExpressionValue(gene, expression))))

    wb.save(savename)

def GeneHistogram(gene, expressiontype):
    # collect all gene expression values
    values = []

    for pt in TCGA_Patients:
        if pt.CheckGeneDict(gene, expressiontype):
            exp = float(pt.GetExpressionValue(gene, expressiontype))
            
            if exp != 0.0:
                values.append( np.log2(float(pt.GetExpressionValue(gene, expressiontype))) )

    plt.hist(values)

    plt.xlabel(f'{gene} expression level')
    plt.ylabel(f'number of samples')
    plt.title(f'{gene} distribution')

    plt.show()
    
# visualize slope data against prediction algorithms too!
# grpreturn is eithe 'ulcerated' or 'non-ulcerated' and returns the significant genes which are downregulated
def GenerateUlcerationHeatmap(expressioncontainer, grpreturn, dx):
    # set up dataframe such that column headers are "ulcerated", "non-ulcerated", "N/A"
    # and row headers are gene names
    global GeneList
    global ProcessedDataBuffer

    clinicalfeaturemap = {'no' : 'non-ulcerated', 'yes' : 'ulcerated'}
    
    ulcerationdata = {'ulcerated' : [np.nan] * len(GeneList), 'non-ulcerated' : [np.nan] * len(GeneList)}
    df = pd.DataFrame(ulcerationdata, index=GeneList.copy())

    # for gene that we know of
    # seggregate the slope values into the dataframe
    for val in GeneList:
        datas = expressioncontainer.SearchForGene(val)
        df.loc[val] = [np.nan, np.nan]# set gene values to NAN
        for data in datas.Data:
            #if data.pval > 0.05:
            #   continue

            if len(data.IValues) < 30:
                continue

            ui = data.GetMetadata('melanoma_ulceration_indicator')

            if ui not in clinicalfeaturemap.keys():
                continue

            if clinicalfeaturemap[ui] == grpreturn:
                if data.pval > 0.05:
                    continue
        
            df.at[val, clinicalfeaturemap[ui]] = data.CorCoeff
            
    # scale to unit variance (SD)
    #scaled_data = StandardScaler().fit_transform(df.transpose())
    #df = pd.DataFrame(scaled_data.transpose(), index = df.index, columns = df.columns)
    # Normalize each value to the non ulcerated and take the log fold change
    for index,row in df.iterrows():
        '''tempU = row['ulcerated']
        tempNU = row['non-ulcerated']
        try:
            row['ulcerated'] = np.log2(tempU / tempNU)
            row['non-ulcerated'] = 1
        except:
            df.drop(index)
        if np.isnan(row['ulcerated']):
            df = df.drop(index)'''
        if row.isnull().any():
            df = df.drop(index)
            continue
        
        if abs(row['ulcerated'] - row['non-ulcerated']) < 0.2:
            df = df.drop(index)
    
    #pdb.set_trace()
    # visualize with heatmap
    # (only ulcerated because we know non ulcerated is 1)
    sns.clustermap(df, z_score=None, standard_scale=None, yticklabels=False)
    plt.title('correlation in ulcerated vs. non-ulcerated groups')
    plt.show()

    if dx == 'pos':
        return df[ df[grpreturn] > 0 ].index.tolist()
    elif dx == 'neg':
        return df[ df[grpreturn] < 0 ].index.tolist()
    else:
        print(f'Invalid value {df} passed to df, should either be pos or neg')
#def GenerateCCFiles(data, stratmetric=None):
#    if stratmetric=None:
        

tm_strings = {'DSS': ['DSS_cr', 'DSS.time.cr'], 'PFI': ['PFI.cr', 'PFI.time.cr']}

'''def SurvivalCurveByFeature(feature, timemetric):
    global TCGA_Patients
    stages = ['stage 0', 'stage i', 'stage ia', 'stage ib', 'stage ic', 'i/ii nos', 'stage ii', 'stage iia', 'stage iib', 'stage iic', 'stage iii', 'stage iiia', 'stage iiib', 'stage iiic', 'stage iv']
    
    # load survival data
    survival_df = pd.read_csv(filedialog.askopenfilename(defaultextension='.csv', filetypes=(("comma seperated values file", "*.csv"),("All Files", "*.*") )))
    
    # build matrix from survival data
    # get SKCM
    skcm_df = survival_df[ survival_df['type'] == 'SKCM']
    
    # get feature data
    feature = [val.GetClinicalFeature(feature) for val in TCGA_Patients if val is not None and val.GetClinicalFeature(feature) is not None]
    
    stranslate = {'Local Disease (i & ii)' : ['stage 0','stage i','stage ia','stage ib','stage ic', 'i/ii nos', 'stage ii', 'stage iia', 'stage iib', 'stage iic'],
                  'Late Stage and Mestastatic (iii & iv)' : ['stage iii', 'stage iiia', 'stage iiib', 'stage iiic', 'stage iv'],
                  'Only Metastatic (iv)' : ['stage iv'],
                  'All Stages' : stages}

    pt_status =  {x : [] for x in stranslate.keys()}
    pfi_time =   {x : [] for x in stranslate.keys()}

    for val in skcm_df['bcr_patient_barcode']:
        pt = GetPatient(val)

        if pt is None:
            continue

        stage = pt.GetClinicalFeature('pathologic_stage')
        
        if stage not in stages:
            continue
        
        for key, value in stranslate.items():
            if stage in value:
                masked = skcm_df[skcm_df['bcr_patient_barcode'] == val]
                tempstatus = masked[tm_strings[timemetric][0]].array[0]
                temptime = masked[tm_strings[timemetric][1]].array[0]

                if tempstatus == '#N/A':
                    continue
                elif temptime == '#N/A' or temptime == 0:
                    continue

                tempexp = pt.GetExpressionValue(gene, expressiontype)
                if tempexp is None:
                    continue
                
                if float(tempexp) > val_high:
                    percentile[key].append('hi')
                elif float(tempexp) < val_low:
                    percentile[key].append('lo')
                else:
                    continue

                if tempstatus == 1:
                    pt_status[key].append(1)
                else:
                    pt_status[key].append(0)

                pfi_time[key].append(float(temptime))
                if np.isnan(temptime): # for some reason a NaN slips through when all of the entries are N/A, temporary workaround
                    percentile[key].pop()
                    pt_status[key].pop()
                    pfi_time[key].pop()

    skcm_dfs = []
    
    for key in stranslate.keys():
        skcm_dfs.append(pd.DataFrame({gene : percentile[key],
                            'Status' : pt_status[key],
                            'Time': pfi_time[key]}))
    
    
    maskhi = skcm_df[gene] == 'hi'
    masklo = skcm_df[gene] == 'lo'
    
    # load into learner of choice
    time_cell, survival_prob_cell, conf_int = kaplan_meier_curve(
        skcm_df['Status'][maskhi], skcm_df['Time'][maskhi])
    
    plt.step(time_cell, survival_prob_cell, where="post", label=f"hi (n = {maskhi.sum()})")
    
    time_cell, survival_prob_cell, conf_int = kaplan_meier_curve(
        skcm_df['Status'][masklo], skcm_df['Time'][masklo])
    
    plt.step(time_cell, survival_prob_cell, where="post", label=f"hi (n = {maskhi.sum()})")


    plt.ylim(0, 1)
    plt.ylabel("est. probability of survival")
    plt.xlabel("time")
    plt.legend(loc="best")
    #fig,axes = plt.subplots(ncols=3,nrows=5)
    colval= GetClosestSquareBounds(len(skcm_dfs))[0]
    rowval = GetClosestSquareBounds(len(skcm_dfs))[1]
    fig,axes = plt.subplots(ncols=colval,nrows=rowval)
    kmf = [KaplanMeierFitter()] * len(skcm_dfs)
    print("=" * 30)
    print(f"SURVIVAL CURVE {percentile_low}-{percentile_high} for {gene}")
    print('='*30)
    if type(axes) is not np.ndarray:
        axes = [axes] # a bit hacky for my taste but should work
    
    for idx,df in enumerate(skcm_dfs):
        hi_df = df[df[gene] == 'hi']
        lo_df = df[df[gene] == 'lo']
        
        if hi_df.empty or lo_df.empty:
            continue

    
        if colval == 1 or rowval == 1:
            kmf[idx].fit(durations=hi_df['Time'], event_observed=hi_df['Status'], label=f'hi n={len(hi_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx])
            kmf[idx].fit(durations=lo_df['Time'], event_observed=lo_df['Status'], label=f'lo n={len(lo_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx])

            results = logrank_test(durations_A=hi_df['Time'], durations_B=lo_df['Time'], event_observed_A=hi_df['Status'], event_observed_B=lo_df['Status'])
            print(f"\nlog rank of group: {list(stranslate.keys())[idx]}")
            results.print_summary()
            print('=' * 30)
            
            axes[idx].set_title(f'Kaplan-Meier Estimates by Group, {list(stranslate.keys())[idx]}')
            axes[idx].set_xlabel('Time')
            axes[idx].set_ylabel('Survival Probability')
            axes[idx].legend()
        else:
            kmf[idx].fit(durations=hi_df['Time'], event_observed=hi_df['Status'], label=f'hi n={len(hi_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx // colval, idx % colval])
            kmf[idx].fit(durations=lo_df['Time'], event_observed=lo_df['Status'], label=f'lo n={len(lo_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx // colval, idx % colval])

            results = logrank_test(durations_A=hi_df['Time'], durations_B=lo_df['Time'], event_observed_A=hi_df['Status'], event_observed_B=lo_df['Status'])
            print(f"\nlog rank of group: {list(stranslate.keys())[idx]}")
            results.print_summary()
            print('=' * 30)
            
            axes[idx // colval, idx % colval].set_title(f'Kaplan-Meier Estimates by Group, {list(stranslate.keys())[idx]}')
            axes[idx // colval, idx % colval].set_xlabel('Time')
            axes[idx // colval, idx % colval].set_ylabel('Survival Probability')
            axes[idx // colval, idx % colval].legend()

    plt.subplots_adjust(hspace=1.5)
    
    fig.show()'''

    
# todo: how to group
# TODO: Fix pickling TCGA_Patients array
# TODO: Histogram!!!

# ptlist is so we can stratify
def SurvivalCurve(percentile_low, percentile_high, timemetric, gene, expressiontype, ptlist=None):
    global TCGA_Patients
    stagefeaturename = 'pathologic_stage'
    stages = ['stage 0', 'stage i', 'stage ia', 'stage ib', 'stage ic', 'i/ii nos', 'stage ii', 'stage iia', 'stage iib', 'stage iic', 'stage iii', 'stage iiia', 'stage iiib', 'stage iiic', 'stage iv']
    
    # load survival data
    survival_df = pd.read_csv(filedialog.askopenfilename(defaultextension='.csv', filetypes=(("comma seperated values file", "*.csv"),("All Files", "*.*") )))
    
    # build matrix from survival data
    # get SKCM
    skcm_df = survival_df[ survival_df['type'] == 'SKCM']
    
    # get mir-196b expression for percentile seggregation
    expression = [float(val.GetExpressionValue(gene, expressiontype)) for val in TCGA_Patients if val is not None and val.GetExpressionValue(gene, expressiontype) is not None]
    
    stranslate = {'Local Disease (i & ii)' : ['stage 0','stage i','stage ia','stage ib','stage ic', 'i/ii nos', 'stage ii', 'stage iia', 'stage iib', 'stage iic'],
                  'Late Stage and Mestastatic (iii & iv)' : ['stage iii', 'stage iiia', 'stage iiib', 'stage iiic', 'stage iv'],
                  'Only Metastatic (iv)' : ['stage iv'],
                  'All Stages' : stages}
    
    #stranslate = {'all stages' : stages}

    #stranslate = {'stage iii' : ['stage iii', 'stage iiia', 'stage iiib', 'stage iiic']}

    #stranslate = {'stage iii' : ['stage iii']}
    
    # REMEMBER TO UPDATE THIS WHEN CHANGING ALGORITHM
    num_graphs = len(stranslate)
    
    # load into data matrix
    percentile = {x : [] for x in stranslate.keys()}
    pt_status =  {x : [] for x in stranslate.keys()}
    pfi_time =   {x : [] for x in stranslate.keys()}

    val_high = np.percentile(expression, percentile_high)
    val_low = np.percentile(expression, percentile_low)    
    
    for val in skcm_df['bcr_patient_barcode']:
        if ptlist is not None:
            if val not in ptlist:
                continue
        
        pt = GetPatient(val)

        if pt is None:
            continue

        stage = pt.GetClinicalFeature('pathologic_stage')
        
        if stage not in stages:
            continue
        
        for key, value in stranslate.items():
            if stage in value:
                masked = skcm_df[skcm_df['bcr_patient_barcode'] == val]
                tempstatus = masked[tm_strings[timemetric][0]].array[0]
                temptime = masked[tm_strings[timemetric][1]].array[0]

                if tempstatus == '#N/A':
                    continue
                elif temptime == '#N/A' or temptime == 0:
                    continue

                tempexp = pt.GetExpressionValue(gene, expressiontype)
                if tempexp is None:
                    continue
                
                if float(tempexp) > val_high:
                    percentile[key].append('hi')
                elif float(tempexp) < val_low:
                    percentile[key].append('lo')
                else:
                    continue

                if tempstatus == 1:
                    pt_status[key].append(1)
                else:
                    pt_status[key].append(0)

                pfi_time[key].append(float(temptime))
                if np.isnan(temptime): # for some reason a NaN slips through when all of the entries are N/A, temporary workaround
                    percentile[key].pop()
                    pt_status[key].pop()
                    pfi_time[key].pop()

    skcm_dfs = []
    
    for key in stranslate.keys():
        skcm_dfs.append(pd.DataFrame({gene : percentile[key],
                            'Status' : pt_status[key],
                            'Time': pfi_time[key]}))
    
    
    '''maskhi = skcm_df[gene] == 'hi'
    masklo = skcm_df[gene] == 'lo'
    
    # load into learner of choice
    time_cell, survival_prob_cell, conf_int = kaplan_meier_curve(
        skcm_df['Status'][maskhi], skcm_df['Time'][maskhi])
    
    plt.step(time_cell, survival_prob_cell, where="post", label=f"hi (n = {maskhi.sum()})")
    
    time_cell, survival_prob_cell, conf_int = kaplan_meier_curve(
        skcm_df['Status'][masklo], skcm_df['Time'][masklo])
    
    plt.step(time_cell, survival_prob_cell, where="post", label=f"hi (n = {maskhi.sum()})")


    plt.ylim(0, 1)
    plt.ylabel("est. probability of survival")
    plt.xlabel("time")
    plt.legend(loc="best")'''
    #fig,axes = plt.subplots(ncols=3,nrows=5)
    colval= GetClosestSquareBounds(len(skcm_dfs))[0]
    rowval = GetClosestSquareBounds(len(skcm_dfs))[1]
    fig,axes = plt.subplots(ncols=colval,nrows=rowval)
    kmf = [KaplanMeierFitter()] * len(skcm_dfs)
    print("=" * 30)
    print(f"SURVIVAL CURVE {percentile_low}-{percentile_high} for {gene}")
    print('='*30)
    if type(axes) is not np.ndarray:
        axes = [axes] # a bit hacky for my taste but should work
    
    for idx,df in enumerate(skcm_dfs):
        hi_df = df[df[gene] == 'hi']
        lo_df = df[df[gene] == 'lo']
        
        if hi_df.empty or lo_df.empty:
            continue

    
        if colval == 1 or rowval == 1:
            kmf[idx].fit(durations=hi_df['Time'], event_observed=hi_df['Status'], label=f'hi n={len(hi_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx])
            kmf[idx].fit(durations=lo_df['Time'], event_observed=lo_df['Status'], label=f'lo n={len(lo_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx])

            results = logrank_test(durations_A=hi_df['Time'], durations_B=lo_df['Time'], event_observed_A=hi_df['Status'], event_observed_B=lo_df['Status'])
            print(f"\nlog rank of group: {list(stranslate.keys())[idx]}")
            results.print_summary()
            print('=' * 30)
            
            axes[idx].set_title(f'Kaplan-Meier Estimates by Group, {list(stranslate.keys())[idx]}')
            axes[idx].set_xlabel('Time')
            axes[idx].set_ylabel('Survival Probability')
            axes[idx].legend()
        else:
            kmf[idx].fit(durations=hi_df['Time'], event_observed=hi_df['Status'], label=f'hi n={len(hi_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx // colval, idx % colval])
            kmf[idx].fit(durations=lo_df['Time'], event_observed=lo_df['Status'], label=f'lo n={len(lo_df.index)}')
            kmf[idx].plot_survival_function(ax=axes[idx // colval, idx % colval])

            results = logrank_test(durations_A=hi_df['Time'], durations_B=lo_df['Time'], event_observed_A=hi_df['Status'], event_observed_B=lo_df['Status'])
            print(f"\nlog rank of group: {list(stranslate.keys())[idx]}")
            results.print_summary()
            print('=' * 30)
            
            axes[idx // colval, idx % colval].set_title(f'Kaplan-Meier Estimates by Group, {list(stranslate.keys())[idx]}')
            axes[idx // colval, idx % colval].set_xlabel('Time')
            axes[idx // colval, idx % colval].set_ylabel('Survival Probability')
            axes[idx // colval, idx % colval].legend()

    plt.subplots_adjust(hspace=1.5)
    
    fig.show()

#def SurvivalGroupsHistogram():
''' Stratify by uniform genetic expression => kstrat 
    Aggregate pt number, multiply by kstrat to get actual expression range
'''



#def PlotSurvivalCurves():


# Pa = A(A^T * A)^(-1) * A^T
# eventually do PCA
#def ProjectPatientData():
    # set up colum vectors of patient gene expression

def GetGeneNamefromTCGAGene(gene):
    return gene[:gene.find('|')]

#TODO: target 
def _psRNATargetGetPredictionSet():
    df = pd.read_csv(filedialog.askopenfilename(defaultextension='.txt', filetypes=(("psRNATarget text output", "*.txt"),("All Files", "*.*"))), sep='\t')
    temp = df['Target_Acc.'].tolist()

    returnval = []
    for val in temp:
        # first word is the gene ID
        genestr = temp[val.find('|')+1:]

        if val in GeneDict.keys():
            returnval.append(GeneDict[val])

    return returnval;

def _MirTarBaseGetPredictionSet():
    df = pd.read_csv( filedialog.askopenfilename(defaultextension='.csv', filetypes=(("comma seperated values file", "*.csv"),("All Files", "*.*"))) )
    temp = df['Target'].tolist()
    global GeneDict
    returnval = []
    for val in temp:
        if type(val) is not str:
            continue

        if val in GeneDict.keys():
            returnval.append(GeneDict[val])

    return returnval

def GenerateExcelOfPredTar(mirna):
    TargetScan = _TargetScanGetPredictionSet(mirna)
    mir_db = _MirDBGetPredictionSet(mirna)
    mirtar = _MirTarBaseGetPredictionSet()
    psrnatar = _psRNATargetGetPredictionSet()

    algorithms = {'TargetScan' : TargetScan, 'mirDB' : mir_db, 'MirTarBase' : mirtar, 'psRNATarget' : psrnatar}
    global GeneDict
    genes = GeneDict.values()

    lists = [{}, {}, {}, {}]
    
    for gene in genes:
        numalgs = 0
        agg = 0
        for key,value in algorithms.items():
            if gene in value:
                numalgs += 1
                agg += value.index(gene)
        
        if not numalgs == 0:
           lists[numalgs-1][gene] = agg

    wb = openpyxl.Workbook()
    for idx,genedict in enumerate(lists):
        genedict = dict(sorted(genedict.items(), key=lambda item: item[1]))
        ws = wb.create_sheet(f"{idx+1} prediction algorithms")
        for key in genedict.keys():
            ws.append([key])

    wb.save(filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=(("excel file", "*.xlsx"),("All Files", "*.*") )))

def GetMirList():
    global GeneList
    
    returnVal = []
    for gene in GeneList:
        if gene[0:3] == 'hsa':
            returnVal.append(gene)

    return returnVal

# do UMAP on miRNA expression with clinical staging and ulceration status
def UMAPStagingMIRNAs(clinicalfeature, genes, expressiontype):
    # colums are genes of interest, then our clinical featuremain
    #mirs = ['hsa-mir-196b', 'hsa-mir-223', 'hsa-mir-135b', 'hsa-mir-363']
    mirs = genes

    # fill up our rows (faster to create temp rows and pass them to the df
    ptdatas = {x : [] for x in (mirs + ['feature'])}
    for pt in TCGA_Patients:
        # append staging (last array)
        stage = pt.GetClinicalFeature(clinicalfeature)
        if stage is None:
            continue
        ptdatas['feature'].append(stage)
        
        for idx,mir in enumerate(mirs):
            val = pt.GetExpressionValue(mir, expressiontype)

            # if we encounter missing data, just remove patient from group
            if val is None:
                ptdatas['feature'].pop()
                for idx2 in range(idx-1):
                    ptdatas[mirs[idx2]].pop()
                break
            
            ptdatas[mir].append(float(val))
            
    pdb.set_trace()

    
    df = pd.DataFrame(ptdatas)
    print(df)

    puredata = df[mirs].values
    scaled_puredata = StandardScaler().fit_transform(puredata)

    reducer = umap.UMAP()
    #transformed = reducer.fit_transform(puredata)
    transformed = reducer.fit_transform(scaled_puredata)

    
    fig,ax = plt.subplots()
    
    # generate feature -> int map for color pallete
    featuremap = {x : idx for idx,x in enumerate(df.feature.value_counts().index.tolist())}
    ax.scatter(transformed[:,0], transformed[:,1],
                c=[sns.color_palette("husl", len(featuremap))[x] for x in df.feature.map(featuremap)])
    ax.set_title(f'UMAP of melanoma miRNA expression colored by {clinicalfeature}')

    fig.show()

# train and test a clasifier
# use component vectors to find genes of interest
# 50% accurate give or take, basically a guess...
# mirlist - list of miRNA's that we are investigating
def LDAmiRUlceration(mirList, expressiontype):
    global TCGA_Patients
    
    dict_targets = {1 : "Ulcerated", 0 : "Non-Ulcerated"}
    
    # fill np array with patient data
    mirvalues = []
    classes = []
    for pt in TCGA_Patients:
        geneexps = []
        if pt.ClinicalFeatures['melanoma_ulceration_indicator'] == 'yes':
            for mir in mirList:
                if pt.CheckGeneDict(mir, expressiontype):
                    geneexps.append(float(pt.GetExpressionValue(mir, expressiontype)))
                else:
                    geneexps.clear()
                    break
            if len(geneexps) == 0: # check if array is empty
                continue
            mirvalues.append(geneexps)
            classes.append(1)
        elif pt.ClinicalFeatures['melanoma_ulceration_indicator'] == 'no':
            for mir in mirList:
                if pt.CheckGeneDict(mir, expressiontype):
                    geneexps.append(float(pt.GetExpressionValue(mir, expressiontype)))
                else:
                    geneexps.clear()
                    break
                
            if len(geneexps) == 0: # check if array is empty
                continue
            mirvalues.append(geneexps)
            classes.append(0)

    # split test and training set
    lda_train_values, lda_test_values, lda_train_classes, lda_test_classes = train_test_split(mirvalues, classes, test_size=0.2, random_state=42)


    # create classifier object and fit model
    clf = LinearDiscriminantAnalysis()
    #clf = QuadraticDiscriminantAnalysis()
    print(f"Values: {len(mirvalues)}")
    print(f"CLasses: {len(classes)}")
    clf.fit(np.array(lda_train_values), np.array(lda_train_classes))


    print(clf.score(np.array(lda_test_values), np.array(lda_test_classes)))

    print(mirList)
    print(clf.coef_)
    print(clf.explained_variance_ratio_)


def CarsonLabGeneNameMiRFix(gene):
    # tack off -Xp|0 at end of name
    if 'p' in gene:
        # last 4 letters
        return gene[:-5]
    else:
        return gene[:-2]
        
    
# gets GEO as a df where each column is expression of a sample
def GetGEOData(geoid, gpl, genenamefix=None):
    gseseries = GEOparse.get_GEO(geo=geoid, destdir='./Downloads')

    # first we need to merge all data so gather list of common genes between all gsms
    # this works, don't try to optimize it or whatever
    templist = next(iter(gseseries.gsms.items())).table['ID_REF'].tolist()
    commongenes = []
    for gene in templist:
        commongenes.append(gene)
        for key,item in gseseries.gsms.items():
            # I know we iterate over one twice
            if gene not in item.table['ID_REF'].tolist():
                commongenes.pop()


    # merge gsms
    

    # fix commongenes
    FinalGenes = []
    if genenamefix is not None:
        for gene in commongenes:
            FinalGenes.append(genenamefix(gene))
'''    
# enrichment analysis on the differential expression of two genes
# each df should be a pandas DataFrame where one column is one patient
def KEGGEnrichmentAnalysis(ad1 : ad.AnnData, ad2 : ad.AnnData):
    # take the mean fold change of patients
    '''

default_pathways = []
# same thing but with TCGA data
#def KEGGEnrichmentAnalysisTCGA(genes, stratificiation_metric, pathwayconfig=None):
    # first get anndata object by populating matrix with expression data
    
    
#def PCAStagingMIRNAs():
    
# dictionary to match webstites to parser algorithms
'''Websites = {'MirDB' : WebOperationGroup(_MirDBSubmitRequest, _MirDBParseResponse) , 'PicTar' : WebOperationGroup(_PicTarSubmitRequest, _PicTarParseResponse),
            'TargetScan' : WebOperationGroup(_TargetScanSubmitRequest, _TargetScanParseResponse),
            'miRWalk' : WebOperationGroup(_MirWalkSubmitRequest, _MirWalkParseResponse)}

# TODO: change this system so no nested calls, that was dumb
class WebsiteInterface:
    def __init__(self, ID):
        self.ID = ID

    def GetPredictionForGene(self, gene):
        # eventually this may need to take metadata into account
        temp = Websites[self.ID]
        return temp.ParseResponse(temp.SubmitRequest)

    @staticmethod
    def GetPredictionForGene(gene, ID):
        return Websites[ID](gene)'''
'''
def CheckDiscrepancy(df, df2, pt_ID):
    mirnas = df['miRNA_ID'].tolist()
    
    xpoints = []
    ypoints = []
    
    for idx,mir in enumerate(mirnas):
        expval = df[df['miRNA_ID'] == mir]['reads_per_million_miRNA_mapped'].iloc[0]
        #expval2 = df2[df['miRNA_ID'] == mir]
        xpoints.append(np.log2(expval))
        try df2[df2[]]:
            xpoints.pop()
            print(mir)
            continue
        else:
            ypoints.append(np.log2(float(pt.GetExpressionValue(mir, 'reads_per_million_miRNA_mapped'))))
    
    print(len(xpoints), len(ypoints))
    mpl.pyplot.plot(xpoints,ypoints)
    mpl.pyplot.show()'''


def QueryGOFiles():
    # check if file is available
    gobasic_url = "https://current.geneontology.org/ontology/go-basic.obo"
    gene2go_url = "https://raw.githubusercontent.com/tanghaibao/goatools/main/data/association"

    gobasic_cachename = "gobasic_temp.obo"
    gene2go_cachename = "gene2go_temp.tsv"

    dirfiles = os.listdir(os.getcwd())

    gobasic = None
    gene2goa = None
    
    if gobasic_cachename not in dirfiles:
        # cache file
        with open(gobasic_cachename, 'w') as file:
            r = requests.get(gobasic_url)
            file.write(r.text)
        
    gobasic = GODag(gobasic_cachename)

    if gene2go_cachename not in dirfiles:
        # cache file
        with open(gene2go_cachename, 'w') as file:
            r = requests.get(gene2go_url)
            file.write(r.text)
        
    gene2goa = IdToGosReader(gene2go_cachename, godag=gobasic)

    return (gobasic, gene2goa)

def GOEnrichmentAnalysis(geneinterests):#, globallist):
    # retrieve go files (ontoloy and gene2go)
    gobasic, gene2goa = QueryGOFiles()

    #fix gene lists, tack of the ID at the end
    for idx,gene in enumerate(geneinterests):
        geneinterests[idx] = gene[:gene.find('|')]

    #for idx,gene in enumerate(globallist):
    #    globallist[idx] = gene[:gene.find('|')]

    temp = geneinterests.copy()
    # perform enrichment
    goeaobj = GOEnrichmentStudy(
    temp,
    gene2goa.get_id2gos(),
    gobasic,
    obsolete='replace',
    methods=['bonferroni', 'fdr_bh'],
    pvalcalc='fisher_scipy_stats')

    return goeaobj

# df is either neg or pos, indicating negative or positive correlation. Siggroup is the significant group
def GetDifferentialCC_UvNU(data, siggroup, df):
    global GeneList

    dfmap = {'neg' : -1, 'pos' : 1}
    factor = 0
    try:
        factor = dfmap[df]
    except:
        print('invalid value was passed for df, it should either be neg or pos!')
        return

    returnVal = []
    
    for gene in GeneList:
        datas = data.SearchForGene(gene)

        # dumb and dangerous fix later
        try:
            Udata = datas.FilterForMetadata('melanoma_ulceration_indicator', 'yes').Data[0]
            NUdata = datas.FilterForMetadata('melanoma_ulceration_indicator', 'no').Data[0]
        except:
            continue

        
        if siggroup == 'ulcerated':
            if Udata.pval < .05 and Udata.CorCoeff * factor > 0 and NUdata.pval > 0.05:
                returnVal.append((gene, Udata.CorCoeff))
        elif siggroup == 'non-ulcerated':
            if NUdata.pval < .05 and NUdata.CorCoeff * factor > 0 and Udata.pval > 0.05:
                returnVal.append((gene, NUdata.CorCoeff))
        else:
            print(f'Invalid value {siggroup} was passed for siggroup, it should be either ulcerated or non-ulcerated!')
            return

    return returnVal

# min_dcc is the minimun change in correlation coefficient between the two groups
def FindGenesToStratifyUlcerationCCForGene(genelist, stratgene, min_dcc):
    expressionlist = [GuessExpressionType(gene) for gene in genelist]
    processeddata = ExpressionDataContainer(ProcessGeneListForClinicalFeature(stratgene, GuessExpressionType(stratgene), genelist, expressionlist, 'melanoma_ulceration_indicator'))

    returnVal = []

    for gene in genelist:
        tempdata = processeddata.SearchForGene(gene)

        try:
            Udata = tempdata.FilterForMetadata('melanoma_ulceration_indicator', 'yes')
            NUdata = tempdata.FilterForMetadata('melanoma_ulceration_indicator', 'no')
        except:
            continue

        if len(Udata.Data[0].IValues) < 40 or len(NUdata.Data[0].IValues) < 40:
            continue
        
        if len(Udata.Data) != 1 or len(NUdata.Data) != 1:
            continue

        if Udata.Data[0].pval > .05 and NUdata.Data[0].pval > .05 and abs(Udata.Data[0].CorCoeff - NUdata.Data[0].CorCoeff) > min_dcc:
            returnVal.append(gene)

    return returnVal

def FindGenesWithHighCC(genelist, stratgene, mincc):
    expressionlist = [GuessExpressionType(gene) for gene in genelist]
    processeddata = ExpressionDataContainer(ProcessGenesForGeneList(stratgene, GuessExpressionType(stratgene), genelist, expressionlist))

    returnVal = []

    for data in processeddata.Data:
        if len(data.IValues) < 40:
            continue

        if abs(data.CorCoeff) > mincc and data.pval < .05:
            returnVal.append((data.DGene, data.CorCoeff))

    return returnVal

def GetCCSortedByPval(datas):
    returnVal = []
    for data in datas:
        if data.pval < .05:
            returnVal.append((data.DGene, data.CorCoeff, data.pval))

    return sorted(returnVal, key=lambda x: x[2])


class BasicMAFEntry:
    def __init__(self, ptID, # patient ID
                 gene, # gene name
                 mutation_type, # mutation (nonsense, missense, etc.
                 nucleotide_poly): # polymorphism (SNP, DEL, etc.)
        self.ptID = ptID
        self.gene_name = gene
        self.mutation_type = mutation_type
        self.nucelotide_poly = nucleotide_poly


class MAFPatient:
    def __init__(self, ptID,
                 entries : list):
        self.ptID = ptID
        self.entries = entries


    def GetMutationTypeOfGene(self, gene):
        l = [x.mutation_type for x in self.entries if x.gene_name == gene]
        return list(set(l))


    def CreateGeneList(self):
        genelist = []
        for entry in self.entries:
            genelist.append(entry.gene_name)
        
        return list(set(genelist))

def GetMAFPatient(ptID, maflist):
    for l in maflist:
        if l.ptID == ptID:
            return l
    return None

    
def LoadMutationsFromFile(filename=None):
    maf_file = ""
    
    if filename is None:
        maf_file = filedialog.askopenfilename(defaultextension='.txt', title="Select File for mutation", filetypes=(("Mutation Annotation Format", "*.maf"),("Text File", "*.txt") ))
    else:
        maf_file = filename

    df = pd.read_csv(maf_file,sep='\t')

    ptID = GetPatientID(df.iloc[0]['Tumor_Sample_Barcode'])
    
    entries = []
    
    for idx,row in df.iterrows():
        if ptID != GetPatientID(row['Tumor_Sample_Barcode']):
            print('Multiple patients from MAF file!')
        
        #create MAF entry
        entries.append(BasicMAFEntry(GetPatientID(row['Tumor_Sample_Barcode']),
                                     row['Hugo_Symbol'] + '|' + str(row['Entrez_Gene_Id']),
                                     row['Variant_Classification'],
                                     row['Variant_Type']))
        
    return MAFPatient(ptID, entries)

def LoadMutationsFromDir():
    maf_dir = filedialog.askdirectory(title="Select MAF directory")

    patients = []

    for root, _, files in os.walk(maf_dir):
        with alive_bar(len(files)) as bar:
            for filename in files:
                # check if this is indeed an maf file
                if filename.find('.maf') == -1:
                    bar()
                    continue
                print(f'Loading: {maf_dir}/{filename}')
                try:
                    patients.append(LoadMutationsFromFile(maf_dir + '/' + filename))
                    bar()
                except:
                    print(f"\033[91mFailed to load MAF file: {maf_dir}/{filename}!\033[0m")
                    bar()
                    
    return patients
    

'''def GetAllPossibleMutationTypesOfGene(patientlist, gene):
    mutation_types = []
    
    for l in patientlist:
        if l.GetMutationTypeOfGene(gene) not in mutation_types:
            mutation_types.append(l.GetMutationTypeOfGene(gene))

    if None in mutation_types:
        print("Multiple mutation types present in one patient!")

    return mutation_types'''
    
def MutationNumber(patientlist, gene, Filter = None, stratification_factor=None, plot : bool = True):
    if not stratification_factor:
        mutationdict = {}
        if Filter is not None:
                mutationdict = {x : 0 for x in Filter}

        mutationdict['No Mutation'] = 0
        
        for l in patientlist:
            
            mutation_type_l = l.GetMutationTypeOfGene(gene)

            if not mutation_type_l:
                mutationdict['No Mutation'] += 1
                continue

            for mutation_type in mutation_type_l:
                if Filter is None:
                    if mutation_type not in mutationdict.keys():
                        mutationdict[mutation_type] = 1
                    else:
                        mutationdict[mutation_type] += 1
                else:
                    if mutation_type in mutationdict.keys():
                        mutationdict[mutation_type] += 1

        if plot is False:
            return mutationdict

        fig,ax = plt.subplots()
        graphstart=0

        for key,item in mutationdict.items():
            ax.bar(gene, item, bottom=graphstart, label=key)
            graphstart += item
            
        ax.set_title("Mutation Count")
        ax.legend()

        plt.show()

        return mutationdict
    else:
        cvalues = GetPossibleFeatureValues(stratification_factor)
        
        mutationdicts = None
        
        if Filter is not None:
            Filter.append('No Mutation')
            mutationdicts = {x : {x : 0 for x in Filter} for x in cvalues}
        else:
            mutationdicts = {x : {'No Mutation' : 0} for x in cvalues}
        
        for l in patientlist:
            pt = GetPatient(l.ptID)
            feature_val = None
            
            if pt is None:
                continue

            feature_val = pt.GetClinicalFeature(stratification_factor)

            if feature_val is None:
                continue
            
            mutation_type_l = l.GetMutationTypeOfGene(gene)

            if not mutation_type_l:
                mutationdicts[feature_val]['No Mutation'] += 1
                continue

            for mutation_type in mutation_type_l:
                if Filter is None:
                    if mutation_type not in mutationdicts[feature_val].keys():
                        mutationdicts[feature_val][mutation_type] = 1
                    else:
                        mutationdicts[feature_val][mutation_type] += 1
                else:
                    if mutation_type in mutationdicts[feature_val].keys():
                         mutationdicts[feature_val][mutation_type] += 1

        if plot is False:
            return mutationdicts

        fig,ax = plt.subplots()

        # get unique list of all mutation types
        mutationtypes = []
        for key0,d in mutationdicts.items():
            mutationtypes += list(d.keys())

        mutationtypes_unique = list(set(mutationtypes))

        tempcolors = randomcolor.RandomColor().generate(count=len(mutationtypes_unique))
        colormap = {x : tempcolors[idx] for idx,x in enumerate(mutationtypes_unique)}
        
        for key0,d in mutationdicts.items():
            tempbottom=0
            for key1,item in d.items():
                ax.bar(key0, item, bottom=tempbottom, color=colormap[key1], label=key1)
                tempbottom += item

        handles, labels = ax.get_legend_handles_labels()
        
        ax.set_title("Mutation Count")
        ax.legend(handles[0:len(mutationtypes_unique)], labels[0:len(mutationtypes_unique)])

        plt.show()

        return mutationdicts
        
Critical_Mutation_List = ['Missense_Mutation', 'Nonsense_Mutation', 'Frame_Shift_Del', 'Frame_Shift_Ins', 'Splice_Site']

def AppendSimpleClinicalFeatureForGeneMutation(genelist : list, maflist, Filter, featurename):
    global TCGA_Patients
    
    for pt in TCGA_Patients:
        maf = GetMAFPatient(pt.ID, maflist)

        if not maf:
            continue

        nomut = True
        for gene in genelist:
            if len(set(maf.GetMutationTypeOfGene(gene)) & set(Filter)) != 0:
                pt.AppendClinicalFeature(featurename, 'yes')
                nomut = False
                break

        if nomut:
            pt.AppendClinicalFeature(featurename, 'no')

def AppendAdvancedClinicalFeatureForGeneMutation(genelist, maflist, Filter, featurename):
    global TCGA_Patients
    
    for pt in TCGA_Patients:
        maf = GetMAFPatient(pt.ID, maflist)

        if not maf:
            continue

        nomut = True
        for gene in genelist:
            if len(set(maf.GetMutationTypeOfGene(gene)) & set(Filter)) != 0:
                pt.AppendClinicalFeature(featurename, maf.GetMutationTypeOfGene(gene))
                nomut = False
                break

        if nomut:
            pt.AppendClinicalFeature(featurename, 'No Mutation')

def GetMAFListGeneList(maflist):
    genelist = []
    for maf in maflist:
        genelist += maf.CreateGeneList()
    
    return list(set(genelist))

def SortUlcerationMutationDifferential(maflist):
    genelist = GetMAFListGeneList(maflist)

    returnVal = []

    with alive_bar(len(genelist)) as bar:
        for gene in genelist:
            tempmut = MutationNumber(maflist, gene, stratification_factor='melanoma_ulceration_indicator', Filter = Critical_Mutation_List, plot = False)

            Ucount = 0
            NUcount = 0

            for key,item in tempmut['no'].items():
                if key == "No Mutation":
                    continue
                NUcount += item

            for key,item in tempmut['yes'].items():
                if key == "No Mutation":
                    continue
                Ucount += item

            returnVal.append((gene, Ucount - NUcount))

            bar()

    return sorted(returnVal, key=lambda x: x[1], reverse=True)

def LoadMethylationData():
    # create anndata and load B values into itmain
    # load datadframe, transpose, convert to np.matrix, load into AnnData
    df = pd.read_csv(filedialog.askopenfilename(title = "Select Firebrowse methylation mean data", defaultextension='.txt', filetypes=(("text file", "*.txt"),("comma seperated values file", "*.csv"),("All Files", "*.*") )), sep='\t')

    df = df[1:]
    df.set_index("Hybridization REF", inplace=True)

    adata = ad.AnnData(df.transpose())
    
    adata.obs_names = [x[:12] for x in adata.obs.index.tolist()]

    # filter duplicates
    duplicates = adata.obs_names.duplicated(keep='first')
    
    return adata[~duplicates,:]

def PlotMethylationDataForClinicalFeature(gene, feature, data, nullgroup, plot=True):
    adata = data[:,gene]
    featurevals = BuildFeatureList(feature)
    
    if not featurevals:
        print(f"clinical feature {feature} is not present within patient data!")
        return

    # filter for featurevals present within the data matrix we were passed
    for k,l in featurevals.items():
        featurevals[k] = list(set(featurevals[k]) & set(adata.obs.index.tolist()))

    
    methyldata = {x : list(adata[featurevals[x],:][:,gene].X.astype(float).flat) for x in featurevals.keys()}

    if not plot:
        ttestvals = {}
        enriched = {}
        for k,v in methyldata.items():
            if k == nullgroup:
                continue

            ttestvals[k] = scipy.stats.ttest_ind(methyldata[nullgroup], v)
            enriched[k] = np.mean(methyldata[nullgroup]) < np.mean(v)
        return (ttestvals, enriched)

    fig,ax = plt.subplots()

    ax.boxplot([v for k,v in methyldata.items()], tick_labels=featurevals.keys())

    ttestvals = {}
    enriched = {}
    for k,v in methyldata.items():
        if k == nullgroup:
            continue
        
        ttestvals[k] = scipy.stats.ttest_ind(methyldata[nullgroup], v)
        enriched[k] = np.mean(methyldata[nullgroup]) < np.mean(v)
    print((ttestvals, enriched))
    plt.show()

    return (ttestvals, enriched)

# df = hi means methylation is higher, lo means lower
def FindMethylationDiffForUlceration(ptlist, df):
    global GeneList
    
    adata = LoadMethylationData()
    adata = adata[ptlist,:]

    returnVal = []
    with alive_bar(len(GeneList)) as bar:
        for gene in GeneList:
            try:
                ttest,enriched = PlotMethylationDataForClinicalFeature(gene[:gene.find('|')], 'melanoma_ulceration_indicator', adata, 'no', plot=False)
                if ttest['yes'].pvalue < .05:
                    if df == 'hi' and enriched:
                        returnVal.append(gene)
                    elif df == 'lo' and not enriched:
                        returnVal.append(gene)
                bar()
            except:
                #print(gene[:gene.find('|')])
                bar()
                continue


    return returnVal
#Restore('patientsave.pickle')
#mrna_df = ImportTCGAData('TCGA_SKCM_mRNA_9.29.23.csv')
#StoreEverything(mrna_df)
#Save('newpatientsave.pickle'
